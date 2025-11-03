import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side
from openpyxl import load_workbook
import pandas as pd
from copy import copy
import argparse

# ======================
# RAW DATA CLEANING AUTOMATION
# ======================
def clean_raw_data(parents_file, kids_file, output_cleaned_file, test_mode=False):
    """Clean raw data files and prepare cleaned kids list with extended months"""
    print("\n" + "="*60)
    print("CLEANING RAW DATA")
    print("="*60)
    
    # Load raw files
    parents_df = pd.read_excel(parents_file, header=None)
    kids_df = pd.read_excel(kids_file, header=None)
    
    # Process kids data structure
    kids_first_rows = kids_df.iloc[:3]
    kids_df = kids_df.iloc[3:].reset_index(drop=True)
    
    # Set column names
    parents_df.columns = [
        "Account_Number", "Booking_Date", "Value_Date", "Transaction_Text",
        "Usage_Purpose", "parent_name", "Account_or_IBAN", "BIC_SWIFT_Code",
        "Amount", "Currency", "Info"
    ]
    
    months_1year = [
        '9', '10', '11', '12', '1', '2', '3',
        '4', '5', '6', '7', '8', '9_next',
        '10_next', '11_next', '12_next', '1_next'
    ]
    
    kids_df.columns = [
        "kid_id", 'kid_name', 'parent_name', *months_1year,
        'class', 'priceOn', 'book_taken', 'nabil_liste'
    ]
    
    backup_kids_df = kids_df.copy()
    
    # Apply test mode if needed
    if test_mode:
        kids_df = kids_df.head(100)
        print(f"🧪 Testing mode: Limited to {len(kids_df)} rows.")
    else:
        # Production mode: stop at first empty kid_id
        if "kid_id" in kids_df.columns:
            stop_index = kids_df["kid_id"].isna().idxmax() if kids_df["kid_id"].isna().any() else None
            if stop_index is not None and stop_index > 0:
                kids_df = kids_df.iloc[:stop_index]
                print(f"✅ Production mode: Stopped at first empty kid_id (row {stop_index}).")
    
    # Parent-kid matching and cleaning
    combined = find_kids_of_parrents(parents_df, kids_df, backup_kids_df)
    
    # Prepare cleaned DataFrame
    cleaned_kids = kids_df.copy()
    cleaned_kids['parent_name'] = cleaned_kids['kid_id'].map(
        combined.set_index('kid_id')['parent_name']
    )
    cleaned_kids['phone_number'] = cleaned_kids['kid_id'].map(
        combined.set_index('kid_id')['phone_number']
    )
    
    # Extend months to 2 years
    months_2years = extend_months_to_2years()
    for month in months_2years:
        if month not in cleaned_kids.columns:
            cleaned_kids[month] = None
    
    # Create final cleaned DataFrame with proper column order
    final_columns = [
        "kid_id", "kid_name", "parent_name", "phone_number",
        *months_2years, "class", "priceOn", "book_taken", "nabil_liste"
    ]
    cleaned_kids = cleaned_kids[final_columns]
    
    # Save cleaned data
    with pd.ExcelWriter(output_cleaned_file) as writer:
        kids_first_rows.to_excel(writer, index=False, header=False)
        cleaned_kids.to_excel(writer, index=False, header=True, startrow=3)
    
    print(f"✅ Cleaned data saved to: {output_cleaned_file}")
    return cleaned_kids

# ======================
# CLEANED DATA PAYMENT AUTOMATION
# ======================
def update_cleaned_data(cleaned_kids_file, parents_file, output_updated_file, test_mode=False):
    """Update payment status on cleaned kids list using latest payments"""
    print("\n" + "="*60)
    print("UPDATING PAYMENT STATUS")
    print("="*60)
    
    # Load files
    parents_df = pd.read_excel(parents_file, header=None)
    wb = load_workbook(cleaned_kids_file)
    ws = wb.active
    
    # Read cleaned kids data
    kids_df = pd.read_excel(cleaned_kids_file, skiprows=3)
    kids_first_rows = pd.read_excel(cleaned_kids_file, nrows=3, header=None)
    
    # Set column names
    parents_df.columns = [
        "Account_Number", "Booking_Date", "Value_Date", "Transaction_Text",
        "Usage_Purpose", "parent_name", "Account_or_IBAN", "BIC_SWIFT_Code",
        "Amount", "Currency", "Info"
    ]
    
    kids_df.columns = [
        "kid_id", "kid_name", "parent_name", "phone_number",
        *extend_months_to_2years(), "class", "priceOn", "book_taken", "nabil_liste"
    ]
    
    # Apply test mode
    if test_mode:
        kids_df = kids_df.head(100)
    
    # Get last updates from cleaned file
    kids_status = get_all_kids_last_updates(cleaned_kids_file)
    
    # Prepare maps
    data_map = get_parent_kid_map(kids_df)
    amount_map = calculate_months_paid(parents_df)
    
    # Calculate payment statuses
    kid_payment_status = calculate_kid_payments(
        data_map, amount_map, 
        {row['kid_name']: {
            'allocated_amount': 0.0,
            'class': row['class'],
            'monthly_fee': get_monthly_fee_for_class(row['class']),
            'parent': row['parent_name']
        } for _, row in kids_df.iterrows()}
    )
    
    # Update Excel with payments
    update_excel_with_payments(
        wb, ws, kids_df, kid_payment_status, kids_status,
        kids_first_rows, output_updated_file
    )
    
    print(f"✅ Updated file saved to: {output_updated_file}")

# ======================
# CORE HELPER FUNCTIONS
# ======================
def extend_months_to_2years():
    """Return list of months covering two full academic years"""
    return [
        '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8',  # Year 1
        '9_next', '10_next', '11_next', '12_next', '1_next', '2_next', 
        '3_next', '4_next', '5_next', '6_next', '7_next', '8_next'  # Year 2
    ]

def find_kids_of_parrents(parents_df, kids_df, backup_kids_df):
    """Match parents to kids and clean parent names"""
    distinct_parents = parents_df['parent_name'].dropna().unique()
    kids_parents = kids_df[['kid_id', 'kid_name', 'parent_name', 'class']].copy()
    
    # Extract phone numbers
    kids_parents['phone_number'] = kids_parents['parent_name'].str.extract(r'\(([^)]*)\)')
    kids_parents['parent_name'] = kids_parents['parent_name'].str.replace(r'\s*\([^\)]*\)', '', regex=True).str.strip()
    
    # Handle missing parent names
    empty_parents = kids_parents[kids_parents['parent_name'].isna()]
    valid_parents = kids_parents[kids_parents['parent_name'].notna()]
    other_distinct_parents = valid_parents['parent_name'].unique()
    
    # Fill missing parents
    for idx, row in empty_parents.iterrows():
        kid_name = row['kid_name']
        matched_parent = None
        
        # Try matching by last name
        for parent in [*other_distinct_parents, *distinct_parents]:
            if not isinstance(parent, str) or not parent.strip():
                continue
            parent_last = parent.split()[-1].lower()
            kid_first = kid_name.split()[0].lower() if kid_name and kid_name.split() else ""
            if parent_last == kid_first:
                matched_parent = parent
                break
        
        # Fallback to backup data
        if not matched_parent and row['kid_id'] in backup_kids_df['kid_id'].values:
            backup_row = backup_kids_df[backup_kids_df['kid_id'] == row['kid_id']].iloc[0]
            matched_parent = backup_row['parent_name']
        
        empty_parents.at[idx, 'parent_name'] = matched_parent if matched_parent else row['parent_name']
    
    # Standardize parent names
    for idx, row in valid_parents.iterrows():
        current_name = str(row['parent_name'])
        for parent in distinct_parents:
            parent_str = str(parent)
            if (parent_str in current_name or current_name in parent_str) and len(parent_str) > 3:
                valid_parents.at[idx, 'parent_name'] = parent
                break
    
    # Combine and sort
    combined = pd.concat([valid_parents, empty_parents], ignore_index=True)
    combined = combined.sort_values(by='kid_id').reset_index(drop=True)
    return combined

def get_parent_kid_map(kids_df):
    """Create mapping of parents to their kids and classes"""
    valid_kids = kids_df[kids_df['parent_name'].notna() & (kids_df['parent_name'].str.strip() != '')]
    multi_kid_parents = valid_kids.groupby('parent_name').filter(lambda x: len(x) >= 2)
    
    parent_map = {}
    for parent, group in multi_kid_parents.groupby('parent_name'):
        parent_map[parent] = dict(zip(group['kid_name'], group['class']))
    return parent_map

def calculate_months_paid(parents_df):
    """Calculate total payments per parent"""
    parents_df['Amount'] = pd.to_numeric(parents_df['Amount'], errors='coerce')
    df_filtered = parents_df.iloc[1:].copy()
    return dict(zip(df_filtered['parent_name'], df_filtered['Amount']))

def get_all_kids_last_updates(file_path):
    """Get last update status for each kid from Excel"""
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    df = pd.read_excel(file_path, skiprows=3)
    
    months = extend_months_to_2years()
    df.columns = [
        "kid_id", "kid_name", "parent_name", "phone_number",
        *months, "class", "priceOn", "book_taken", "nabil_liste"
    ]
    
    results = []
    for idx, row in df.iterrows():
        if pd.isna(row['kid_name']):
            continue
            
        excel_row = idx + 4  # Account for header rows
        last_update = get_last_kid_update(sheet, df, excel_row)
        
        results.append({
            "kid_id": row["kid_id"],
            "kid_name": row['kid_name'],
            "parent_name": row["parent_name"],
            "last_month": last_update["month"],
            "last_text": last_update["text"],
            "last_color": last_update["color"]
        })
    
    return pd.DataFrame(results)

def get_last_kid_update(sheet, df, row_idx):
    """Get last non-red update for a kid"""
    months = extend_months_to_2years()
    last_update = {"month": None, "text": None, "color": "FF595959"}
    
    for month in reversed(months):
        col_idx = df.columns.get_loc(month) + 1
        cell = sheet.cell(row=row_idx, column=col_idx)
        text = str(cell.value).strip() if cell.value else ""
        
        # Get cell color
        try:
            color = cell.fill.start_color.rgb or "FF595959"
        except:
            color = "FF595959"
        
        # Skip red cells
        if color == "FFFF0000":
            continue
        
        if text or (color not in ["00000000", "None", ""]):
            last_update = {"month": month, "text": text, "color": color}
            break
    
    return last_update

def get_monthly_fee_for_class(class_name):
    """Determine monthly fee based on class"""
    A5_names = ["A5","A6","A7","A8","A9","A10","A11","A12","G2"]
    B0_names = ["B0","B1","B2","B3","G1"]
    
    class_clean = str(class_name).strip()
    if class_clean in A5_names:
        return 25.0
    if class_clean in B0_names:
        return 15.0
    return 25.0  # Default

def determine_status_and_color(months_paid, monthly_fee, allocated_amount, class_name):
    """Determine payment status text and color"""
    class_clean = str(class_name).strip()
    
    if monthly_fee <= 0:
        return "Not yet registered", "FF595959"
    
    if allocated_amount == 0:
        return "Nothing paid.", "FFFF0000"
    
    if months_paid >= 1.0:
        return "Fully paid.", "FF92D050"
    
    if class_clean == 'A5' and abs(allocated_amount - 15) < 0.01:
        return "G1 and G2 paid €15 instead of €25.", "FFFFFF00"
    
    if class_clean == 'A5' and allocated_amount in [10, 15, 20]:
        return "Transfers only €10, €15, or €20 instead of €25.", "FFC65911"
    if class_clean == 'B0' and allocated_amount == 10:
        return "Transfers only €10 instead of €15.", "FFC65911"
    
    return f"Partial payment: {allocated_amount:.2f}€ ({months_paid:.2f} months)", "FFFFC000"

def calculate_kid_payments(data_map, amount_map, kids_status):
    """Calculate payment allocation for each kid"""
    kid_payment_status = {}
    prior_alloc = {kid: info['allocated_amount'] for kid, info in kids_status.items()}
    prior_parent_total = {}
    
    for kid, info in kids_status.items():
        parent = info['parent']
        prior_parent_total[parent] = prior_parent_total.get(parent, 0) + prior_alloc.get(kid, 0)
    
    for parent, kids in data_map.items():
        new_payment = float(amount_map.get(parent, 0.0))
        prior_total = prior_parent_total.get(parent, 0.0)
        total_amount = prior_total + new_payment
        
        # Calculate monthly fees
        kid_fees = {}
        total_fee = 0
        for kid, cls in kids.items():
            fee = get_monthly_fee_for_class(cls)
            kid_fees[kid] = fee
            total_fee += fee
        
        if total_fee <= 0:
            for kid, cls in kids.items():
                kid_payment_status[kid] = {
                    'parent': parent,
                    'class': cls,
                    'monthly_fee': 0,
                    'allocated_amount': 0,
                    'months_paid': 0,
                    'status': "Not yet registered",
                    'color': "FF595959"
                }
            continue
        
        # Allocate payments
        full_months = int(total_amount // total_fee)
        remainder = total_amount - (full_months * total_fee)
        
        allocations = {}
        kid_list = list(kids.items())
        for i, (kid, cls) in enumerate(kid_list):
            base_alloc = full_months * kid_fees[kid]
            allocations[kid] = base_alloc + (remainder if i == 0 else 0)
        
        # Build status
        for kid, cls in kids.items():
            alloc = allocations[kid]
            fee = kid_fees[kid]
            months = alloc / fee if fee > 0 else 0
            
            status, color = determine_status_and_color(months, fee, alloc, cls)
            extras = alloc - (int(months) * fee) if fee > 0 else 0
            
            kid_payment_status[kid] = {
                'parent': parent,
                'class': cls.strip(),
                'monthly_fee': round(fee, 2),
                'allocated_amount': round(alloc, 2),
                'months_paid': round(months, 2),
                'status': status,
                'color': color,
                'extras': round(extras, 2),
                'extras_color': "#ffc000" if extras > 0.01 else "#92d050"
            }
    
    return kid_payment_status

def update_excel_with_payments(wb, ws, kids_df, kid_payment_status, kids_status, 
                             kids_first_rows, output_file):
    """Update Excel file with payment statuses while preserving formatting"""
    months = extend_months_to_2years()
    month_start_col = 5  # After kid_id, kid_name, parent_name, phone_number
    
    # Create status dictionary
    status_dict = {row['kid_name']: row for _, row in kids_status.iterrows()}
    
    # Process each kid
    for idx, row in kids_df.iterrows():
        kid_name = row['kid_name']
        if pd.isna(kid_name):
            continue
        
        excel_row = idx + 4  # Account for header rows
        payment_info = kid_payment_status.get(kid_name, {})
        last_status = status_dict.get(kid_name, {})
        
        if not payment_info or not last_status:
            continue
        
        # Get last update position
        try:
            last_idx = months.index(last_status['last_month']) if last_status.get('last_month') else -1
        except ValueError:
            last_idx = -1
        
        # Determine update range
        months_paid = payment_info.get('months_paid', 0)
        full_months = int(months_paid)
        extras = payment_info.get('extras', 0)
        base_color = payment_info.get('color', 'FF595959')
        extras_color = payment_info.get('extras_color', '#92d050')
        
        # Update month cells
        for i, month in enumerate(months):
            col = month_start_col + i
            cell = ws.cell(row=excel_row, column=col)
            
            if i <= last_idx:  # Preserve existing data
                continue
            
            if i <= last_idx + full_months:  # Full month payments
                cell.value = int(payment_info['monthly_fee']) if payment_info['monthly_fee'].is_integer() else payment_info['monthly_fee']
                cell.fill = PatternFill(start_color=base_color[1:], end_color=base_color[1:], fill_type="solid")
            elif i == last_idx + full_months + 1 and extras > 0:  # Extras
                cell.value = extras
                cell.fill = PatternFill(start_color=extras_color[1:], end_color=extras_color[1:], fill_type="solid")
            elif i == last_idx + full_months + 1:  # First unpaid month
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Save updated file
    wb.save(output_file)

# ======================
# MAIN EXECUTION
# ======================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Kids payment automation system')
    parser.add_argument('--mode', choices=['raw', 'cleaned', 'both'], default='both',
                        help='Processing mode: raw (clean data only), cleaned (update payments only), both (full process)')
    parser.add_argument('--test', action='store_true', help='Run in test mode (limited rows)')
    args = parser.parse_args()
    
    print(f"🚀 Starting automation in {args.mode.upper()} mode with test_mode={args.test}")
    
    if args.mode in ['raw', 'both']:
        clean_raw_data(
            parents_file="parents_payments.xlsx",
            kids_file="kids_list.xlsx",
            output_cleaned_file="kids_list_cleaned.xlsx",
            test_mode=args.test
        )
    
    if args.mode in ['cleaned', 'both']:
        update_cleaned_data(
            cleaned_kids_file="kids_list_cleaned.xlsx" if args.mode == 'cleaned' else "kids_list_cleaned.xlsx",
            parents_file="parents_payments.xlsx",
            output_updated_file="kids_list_updated.xlsx",
            test_mode=args.test
        )
    
    print("\n🎉 Automation completed successfully!")