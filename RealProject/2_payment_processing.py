"""
Payment Processing Script
==========================
This script takes CLEANED Excel files and updates payment status for all kids.

Input files (must be cleaned first):
- parents_payments_cleaned.xlsx
- kids_list_cleaned.xlsx

Output files:
- kids_list_updated.xlsx (with new payment status)
- payment_status_report.xlsx (detailed report)
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Color
from copy import copy

# ============================================================================
# CONFIGURATION
# ============================================================================

INPUT_PARENTS_CLEANED = "parents_payments_cleaned.xlsx"
INPUT_KIDS_CLEANED = "kids_list_cleaned.xlsx"

OUTPUT_KIDS_UPDATED = "kids_list_updated.xlsx"
OUTPUT_PAYMENT_REPORT = "payment_status_report.xlsx"

# Monthly fees by class category
MONTHLY_FEE_A = 25.0  # A5, A6, A7, etc.
MONTHLY_FEE_B = 15.0  # B0, B1, B2, etc.

A5_CLASSES = ["A5", "A6", "A7", "A8", "A9", "A10", "A11", "A12", "G2"]
B0_CLASSES = ["B0", "B1", "B2", "B3", "G1", "G2"]

# All months (2 years)
MONTHS_2_YEARS = [
    '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8',  # Year 1
    '9_next', '10_next', '11_next', '12_next', '1_next', '2_next', 
    '3_next', '4_next', '5_next', '6_next', '7_next', '8_next'  # Year 2
]

# Status colors
STATUS_COLORS = {
    "Not yet registered": "FF595959",
    "Nothing paid": "FFFF0000",
    "Fully paid": "FF92D050",
    "G1/G2 underpaid": "FFFFFF00",
    "Underpaid": "FFC65911",
    "Partial payment": "FFFFC000"
}

# ============================================================================
# STEP 1: LOAD CLEANED DATA
# ============================================================================

def load_cleaned_data():
    """Load cleaned Excel files"""
    print("📂 Loading cleaned data files...")
    
    parents_df = pd.read_excel(INPUT_PARENTS_CLEANED)
    kids_df = pd.read_excel(INPUT_KIDS_CLEANED)
    
    print(f"✅ Loaded {len(parents_df)} parent payments")
    print(f"✅ Loaded {len(kids_df)} kids")
    
    return parents_df, kids_df


# ============================================================================
# STEP 2: ANALYZE CURRENT PAYMENT STATUS
# ============================================================================

def get_monthly_fee(class_name):
    """Get monthly fee based on class"""
    if class_name in A5_CLASSES:
        return MONTHLY_FEE_A
    elif class_name in B0_CLASSES:
        return MONTHLY_FEE_B
    return MONTHLY_FEE_A  # default


def get_last_kid_update(sheet, df, row_idx):
    """Get the last payment update for a kid"""
    last_update = {"month": None, "text": None, "color": None}
    
    for month in reversed(MONTHS_2_YEARS):
        if month not in df.columns:
            continue
        
        col_idx = df.columns.get_loc(month) + 1
        cell = sheet.cell(row=row_idx, column=col_idx)
        text = str(cell.value).strip() if cell.value else ""
        
        # Get cell color
        color = None
        try:
            fill_color = cell.fill.start_color
            if isinstance(fill_color, Color):
                if fill_color.rgb and isinstance(fill_color.rgb, str):
                    color = fill_color.rgb
                elif fill_color.indexed is not None:
                    color = str(fill_color.indexed)
            else:
                color = None
        except:
            color = "FF595959"
        
        # Skip red cells (unpaid markers)
        if color == "FFFF0000":
            continue
        
        # Found meaningful content
        if text or (color and color not in ["00000000", "None", ""]):
            last_update = {"month": month, "text": text or None, "color": color}
            break
    
    return last_update


def analyze_current_status(file_path):
    """Analyze current payment status from Excel"""
    print("\n🔍 Analyzing current payment status...")
    
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    df = pd.read_excel(file_path)
    results = []
    
    for index, row in df.iterrows():
        kid_name = row['kid_name']
        if pd.isna(kid_name):
            continue
        
        excel_row_idx = index + 2  # +2 for header
        update = get_last_kid_update(sheet, df, excel_row_idx)
        
        results.append({
            "kid_id": row["kid_id"],
            "kid_name": kid_name,
            "parent_name": row["parent_name"],
            "class": row["class"],
            "last_month": update["month"],
            "last_text": update["text"],
            "last_color": update["color"]
        })
    
    status_df = pd.DataFrame(results)
    print(f"✅ Analyzed status for {len(status_df)} kids")
    
    return status_df


# ============================================================================
# STEP 3: BUILD PARENT-KID MAPPING
# ============================================================================

def build_parent_kid_map(kids_df):
    """Build map of parents to their kids"""
    print("\n👨‍👩‍👧‍👦 Building parent-kid relationships...")
    
    parent_kid_map = {}
    
    df_valid = kids_df[kids_df['parent_name'].notna() & (kids_df['parent_name'].str.strip() != '')]
    parent_groups = df_valid.groupby('parent_name').filter(lambda x: len(x) >= 2)
    
    for parent, group in parent_groups.groupby('parent_name'):
        parent_kid_map[parent] = dict(zip(group['kid_name'], group['class']))
    
    print(f"✅ Found {len(parent_kid_map)} parents with multiple kids")
    
    return parent_kid_map


def build_amount_map(parents_df):
    """Build map of parents to payment amounts"""
    print("\n💰 Building payment amounts map...")
    
    parents_df['Amount'] = pd.to_numeric(parents_df['Amount'], errors='coerce')
    amount_map = dict(zip(parents_df['parent_name'], parents_df['Amount']))
    
    total_amount = sum(v for v in amount_map.values() if v and v > 0)
    print(f"✅ Total payments: €{total_amount:.2f} from {len(amount_map)} parents")
    
    return amount_map


# ============================================================================
# STEP 4: CALCULATE NEW PAYMENT STATUS
# ============================================================================

def determine_status_and_color(months_paid, monthly_fee, allocated_amount, class_name):
    """Determine payment status and color"""
    if monthly_fee <= 0:
        return "Not yet registered", STATUS_COLORS["Not yet registered"]
    
    if allocated_amount == 0:
        return "Nothing paid", STATUS_COLORS["Nothing paid"]
    
    if months_paid >= 1.0:
        return "Fully paid", STATUS_COLORS["Fully paid"]
    
    # Special cases
    if class_name.strip() == 'A5' and abs(allocated_amount - 15) < 0.01:
        return "G1/G2 paid €15 instead of €25", STATUS_COLORS["G1/G2 underpaid"]
    
    if class_name.strip() == 'A5' and allocated_amount in [10, 15, 20]:
        return f"Underpaid: €{allocated_amount} instead of €25", STATUS_COLORS["Underpaid"]
    
    if class_name.strip() == 'B0' and allocated_amount == 10:
        return f"Underpaid: €10 instead of €15", STATUS_COLORS["Underpaid"]
    
    return f"Partial: €{allocated_amount:.2f} ({months_paid:.2f} months)", STATUS_COLORS["Partial payment"]


def calculate_kid_payments(parent_kid_map, amount_map, kids_status_df):
    """Calculate payment status for all kids"""
    print("\n🧮 Calculating payment allocations...")
    
    kid_payment_status = {}
    
    # Build prior allocation from status
    prior_kid_alloc = {}
    prior_parent_total = {}
    
    for _, row in kids_status_df.iterrows():
        kid_name = row['kid_name']
        last_text = row.get('last_text', '')
        
        # Try to extract prior payment from last_text
        try:
            if last_text and isinstance(last_text, str):
                alloc = float(''.join(filter(str.isdigit, last_text.split()[0])))
            else:
                alloc = 0.0
        except:
            alloc = 0.0
        
        prior_kid_alloc[kid_name] = alloc
        
        parent = row['parent_name']
        if parent:
            prior_parent_total[parent] = prior_parent_total.get(parent, 0.0) + alloc
    
    # Process each parent
    for parent, kids in parent_kid_map.items():
        new_payment = float(amount_map.get(parent, 0.0))
        prior_total = prior_parent_total.get(parent, 0.0)
        total_effective = prior_total + new_payment
        
        print(f"\n  {parent}: €{new_payment} new + €{prior_total} prior = €{total_effective} total")
        
        # Get monthly fees
        kid_list = list(kids.items())
        kid_fees = {kid: get_monthly_fee(cls) for kid, cls in kid_list}
        total_monthly_fee = sum(kid_fees.values())
        
        if total_monthly_fee <= 0:
            for kid_name, class_name in kid_list:
                kid_payment_status[kid_name] = {
                    'parent': parent,
                    'class': class_name,
                    'monthly_fee': 0.0,
                    'allocated_amount': 0.0,
                    'months_paid': 0.0,
                    'status': "Not yet registered",
                    'color': STATUS_COLORS["Not yet registered"],
                    'extras': 0.0
                }
            continue
        
        # Allocate payment
        full_months_total = int(total_effective // total_monthly_fee)
        remainder = total_effective - (full_months_total * total_monthly_fee)
        
        # Distribute to kids
        for i, (kid_name, class_name) in enumerate(kid_list):
            monthly_fee = kid_fees[kid_name]
            base_alloc = full_months_total * monthly_fee
            
            # First kid gets remainder
            allocated = base_alloc + (remainder if i == 0 else 0)
            months_paid = allocated / monthly_fee if monthly_fee > 0 else 0.0
            
            full_months_kid = int(allocated // monthly_fee)
            extras = allocated - (full_months_kid * monthly_fee)
            
            status_msg, color = determine_status_and_color(months_paid, monthly_fee, allocated, class_name)
            
            kid_payment_status[kid_name] = {
                'parent': parent,
                'class': class_name,
                'monthly_fee': round(monthly_fee, 2),
                'allocated_amount': round(allocated, 2),
                'months_paid': round(months_paid, 2),
                'status': status_msg,
                'color': color,
                'extras': round(extras, 2)
            }
            
            print(f"    → {kid_name}: €{allocated:.2f} = {months_paid:.2f} months")
    
    print(f"\n✅ Calculated payments for {len(kid_payment_status)} kids")
    return kid_payment_status


# ============================================================================
# STEP 5: UPDATE EXCEL FILE
# ============================================================================

def copy_cell_format(source_cell, target_cell):
    """Copy formatting from source to target cell"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)


def update_excel_with_payments(kids_df, kid_payment_status, kids_status_df, output_file):
    """Update Excel file with new payment status"""
    print("\n📝 Updating Excel file...")
    
    # Load workbook
    wb = load_workbook(INPUT_KIDS_CLEANED)
    ws = wb.active
    
    # Get reference cell for styling
    month_start_col = 4
    reference_cell = ws.cell(row=2, column=month_start_col)
    
    # Convert status to dict for lookup
    status_dict = {}
    for _, row in kids_status_df.iterrows():
        status_dict[row['kid_name']] = {
            'last_month': row['last_month'],
            'last_color': (row.get('last_color') or "").upper().replace("#", "")
        }
    
    # Process each kid
    start_row = 2
    for idx, kid_row in kids_df.iterrows():
        excel_row = start_row + idx
        kid_name = kid_row['kid_name']
        
        if pd.isna(kid_name):
            continue
        
        payment_info = kid_payment_status.get(kid_name, {})
        if not payment_info:
            continue
        
        last_status = status_dict.get(kid_name, {})
        last_month = last_status.get('last_month')
        last_color = last_status.get('last_color', "")
        
        is_not_registered = last_color in ["FF595959", "595959"]
        
        # Skip if not registered and no payment
        if is_not_registered and payment_info.get('allocated_amount', 0) == 0:
            continue
        
        # Get month index
        try:
            last_month_idx = MONTHS_2_YEARS.index(last_month) if last_month else -1
        except ValueError:
            last_month_idx = -1
        
        # Get payment details
        months_paid = payment_info.get('months_paid', 0.0)
        full_months = int(months_paid)
        monthly_fee = payment_info.get('monthly_fee', 0.0)
        new_color = payment_info.get('color', 'FF595959')
        extras = payment_info.get('extras', 0.0)
        
        # Fill month columns
        for i, month in enumerate(MONTHS_2_YEARS):
            col_idx = month_start_col + i
            cell = ws.cell(row=excel_row, column=col_idx)
            
            # Keep old data
            if i <= last_month_idx:
                continue
            
            # Fill paid months
            elif i <= last_month_idx + full_months:
                cell.value = int(monthly_fee) if monthly_fee == int(monthly_fee) else monthly_fee
                copy_cell_format(reference_cell, cell)
                cell.fill = PatternFill(start_color=new_color.replace("#", ""),
                                       end_color=new_color.replace("#", ""),
                                       fill_type="solid")
            
            # Extras month
            elif i == last_month_idx + full_months + 1 and extras > 0:
                cell.value = extras if extras != int(extras) else int(extras)
                copy_cell_format(reference_cell, cell)
                cell.fill = PatternFill(start_color="FFFFC000",
                                       end_color="FFFFC000",
                                       fill_type="solid")
            
            # First unpaid month (red marker)
            elif i == last_month_idx + full_months + 1 and extras == 0:
                cell.fill = PatternFill(start_color="FFFF0000",
                                       end_color="FFFF0000",
                                       fill_type="solid")
        
        print(f"  ✓ {kid_name}: {full_months} months, €{extras:.2f} extra")
    
    # Save
    wb.save(output_file)
    print(f"\n✅ Saved: {output_file}")


def generate_payment_report(kid_payment_status):
    """Generate detailed payment report"""
    print("\n📊 Generating payment report...")
    
    report_data = []
    for kid_name, info in kid_payment_status.items():
        report_data.append({
            'Kid Name': kid_name,
            'Parent': info['parent'],
            'Class': info['class'],
            'Monthly Fee': info['monthly_fee'],
            'Allocated Amount': info['allocated_amount'],
            'Months Paid': info['months_paid'],
            'Extras': info['extras'],
            'Status': info['status']
        })
    
    report_df = pd.DataFrame(report_data)
    report_df.to_excel(OUTPUT_PAYMENT_REPORT, index=False)
    
    print(f"✅ Saved report: {OUTPUT_PAYMENT_REPORT}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    print("="*70)
    print("PAYMENT PROCESSING SCRIPT")
    print("="*70)
    
    # Load cleaned data
    parents_df, kids_df = load_cleaned_data()
    
    # Analyze current status
    kids_status_df = analyze_current_status(INPUT_KIDS_CLEANED)
    
    # Build maps
    parent_kid_map = build_parent_kid_map(kids_df)
    amount_map = build_amount_map(parents_df)
    
    # Calculate payments
    kid_payment_status = calculate_kid_payments(parent_kid_map, amount_map, kids_status_df)
    
    # Update Excel
    update_excel_with_payments(kids_df, kid_payment_status, kids_status_df, OUTPUT_KIDS_UPDATED)
    
    # Generate report
    generate_payment_report(kid_payment_status)
    
    print("\n" + "="*70)
    print("✅ PAYMENT PROCESSING COMPLETED!")
    print("="*70)
    print(f"\nCheck these files:")
    print(f"  • {OUTPUT_KIDS_UPDATED}")
    print(f"  • {OUTPUT_PAYMENT_REPORT}")


if __name__ == "__main__":
    main()