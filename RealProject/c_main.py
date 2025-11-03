import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color ,   Border, Side
from openpyxl import load_workbook
import pandas as pd

from copy import copy
# Load the Excel file (or CSV)
parents_df = pd.read_excel("parents_payments.xlsx" , header=None,)
kids_df = pd.read_excel("kids_list.xlsx",header=None,)
# Split the first 3 rows
kids_first_rows = kids_df.iloc[:3]   # first 3 rows
kids_df = kids_df.iloc[3:]    # all rows after the first 3
kids_df = kids_df.reset_index(drop=True)

kids_last_rows = kids_df
# refinning kids_df , parents_df columns names
parents_df.columns = [
    "Account_Number",            # Auftragskonto
    "Booking_Date",              # Buchungstag
    "Value_Date",                # Valutadatum
    "Transaction_Text",          # Buchungstext
    "Usage_Purpose",             # Verwendungszweck
    "parent_name",        # Beguenstigter/Zahlungspflichtiger
    "Account_or_IBAN",           # Kontonummer/IBAN
    "BIC_SWIFT_Code",            # BIC (SWIFT-Code)
    "Amount",                    # Betrag
    "Currency",                  # Waehrung
    "Info"                       # Info
]
wb = load_workbook("kids_list.xlsx")
ws = wb.active
last_colomn = ws.max_column

months =[]
if last_colomn > 22:
    months = [
        '9', '10', '11', '12', '1', '2', '3',
        '4', '5', '6', '7', '8','9_next',
        '10_next', '11_next', '12_next', '1_next'
        ]
else:
    months = [
        '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8',  # Year 1
        '9_next', '10_next', '11_next', '12_next', '1_next', '2_next', 
        '3_next', '4_next', '5_next', '6_next', '7_next', '8_next'  # Year 2
    ]

kids_df.columns = [
    "kid_id",'kid_name', 'parent_name', *months,
    'class' , 'priceOn','book_taken','nabil_liste'
    ]

backup_kids_df = kids_df.copy()



# print("Data loaded successfully.")

# print("Showing data samples:")
# print("Parents Payments Data:")
# # print(parents_df.head())
# print("\nKids Full Names Data:")
# # print(kids_df.head())

# Choose mode: "test" or "prod"
mode = "prod"  # change to "test" when testing

if mode == "test":
    # Limit to first 100 rows
    kids_df = kids_df.head(100)
    print(f"🧪 Testing mode: Limited to {len(kids_df)} rows.")
else:
    # Production mode: stop at first empty kid_id
    if "kid_id" in kids_df.columns:
        # Find first row where kid_id is empty
        stop_index = kids_df["kid_id"].isna().idxmax() if kids_df["kid_id"].isna().any() else None
        
        if stop_index is not None and stop_index > 0:
            kids_last_rows = kids_df.iloc[stop_index:]
            kids_df = kids_df.iloc[:stop_index]
            print(f"✅ Production mode: Stopped at first empty kid_id (row {stop_index}).")
        else:
            print("✅ Production mode: No missing kid_id found. Using all rows.")
    else:
        print("⚠️ Column 'kid_id' not found in kids_df!")



# all of them
# find parents names in kids full names
def find_kids_of_parrents(parents_df, kids_df):
    #step 1: find distinct parents names
    # 1. Distinct parents from parents_df
    distinct_parents = parents_df['parent_name'].dropna().unique()

    # 2. Distinct kid-parent pairs from kids_df
    # Example DataFrame
    kids_parents_from_kids = kids_df[['kid_id','kid_name', 'parent_name' , 'class']]

    # Extract content inside parentheses to new column
    kids_parents_from_kids['phone_number'] = kids_parents_from_kids['parent_name'].str.extract(r'\(([^)]*)\)')

    # Remove parentheses and content from parent_name
    kids_parents_from_kids['parent_name'] = kids_parents_from_kids['parent_name'].str.replace(r'\s*\([^\)]*\)', '', regex=True)

    # Remove leading/trailing spaces
    kids_parents_from_kids['parent_name'] = kids_parents_from_kids['parent_name'].str.strip()
    kids_parents_from_kids.to_excel("kids_parents_from_kids_debug21.xlsx", index=False)

    # Drop empty strings
    empty_kids_parents_from_kids = kids_parents_from_kids[kids_parents_from_kids['parent_name'].isna()]
    kids_parents_from_kids = kids_parents_from_kids[kids_parents_from_kids['parent_name'].notna()]
    other_distinct_parents = kids_parents_from_kids['parent_name'].dropna().unique()
    # print(f"Distinct parents from parents list: {distinct_parents}")
    
    # completing empty parent names from distinct parents
    # Loop over the rows with missing parent names
    for index, row in empty_kids_parents_from_kids.iterrows():
        kid_name = row['kid_name']
        matched_parent = None

        # First, check other distinct parents
        for parent in other_distinct_parents:
            if not isinstance(parent, str) or not parent.strip():
                continue  # skip empty or non-string parents
            last_name_parent = parent.split()[-1]
            if kid_name and kid_name.split() and last_name_parent.lower() == kid_name.split()[0].lower():
                matched_parent = parent
                break

        # If no match found, check main distinct parents
        if not matched_parent:
            for parent in distinct_parents:
                if not isinstance(parent, str) or not parent.strip():
                    continue
                last_name_parent = parent.split()[-1]
                if kid_name and kid_name.split() and last_name_parent.lower() == kid_name.split()[0].lower():
                    matched_parent = parent
                    break

        # Update the main DataFrame if a match was found
        if matched_parent:
            empty_kids_parents_from_kids.at[index, 'parent_name'] = matched_parent
            print(f"Completed missing parent name for kid '{kid_name}' with parent '{matched_parent}'")
        else:
            print(f"No matching parent found for kid '{kid_name}' with missing parent name will put the original")
            # Make sure both DataFrames have 'kid_id' column
            for kid_id in empty_kids_parents_from_kids['kid_id']:
                # Find the row in empty_kids_parents_from_kids
                idx = empty_kids_parents_from_kids.index[empty_kids_parents_from_kids['kid_id'] == kid_id][0]
                
                # Only update if parent_name is empty
                if pd.isna(empty_kids_parents_from_kids.at[idx, 'parent_name']) or empty_kids_parents_from_kids.at[idx, 'parent_name'] == '':
                    # Get the parent_name from backup_kids_df using kid_id
                    parent_name = backup_kids_df.loc[backup_kids_df['kid_id'] == kid_id, 'parent_name'].values
                    if len(parent_name) > 0:
                        empty_kids_parents_from_kids.at[idx, 'parent_name'] = str(parent_name[0])


    empty_kids_parents_from_kids.to_excel("kids_parents_from_kids_debug23.xlsx", index=False)
    # After filling missing names, sort the DataFrame by kid_id
    # kids_parents_from_kids = kids_parents_from_kids.sort_values(by='kid_id').reset_index(drop=True)
    # replaceing parents names in  kids_parents_from_kids with distinct parents names if match found or in
    for index, row in kids_parents_from_kids.iterrows():
        kid_name = row['kid_name']
        current_parent_name = row['parent_name']
        matched_parent = None
        for parent in distinct_parents:
            try:
                parent_str = str(parent)
                current_parent_name_str = str(current_parent_name)
                if parent_str in current_parent_name_str or current_parent_name_str in parent_str and len(parent_str) > 3:
                    matched_parent = parent
                    break
            except TypeError:
                print("TypeError encountered!")
                print(f"parent: {parent} ({type(parent)})")
                print(f"current_parent_name: {current_parent_name} ({type(current_parent_name)})")
                continue


        if matched_parent:
            kids_parents_from_kids.at[index, 'parent_name'] = matched_parent
            print(f"Replaced parent name for kid '{kid_name}' with parent New:'{matched_parent}' , Old:'{current_parent_name}'")
        else:
            print(f"No matching parent found to replace for kid '{kid_name}' with parent name '{current_parent_name}'")

    # export excel file for kids_parents_from_kids for debugging
    # Concatenate the two DataFrames
    combined = pd.concat([kids_parents_from_kids, empty_kids_parents_from_kids], ignore_index=True)

    # Sort by 'kid_id'
    combined = combined.sort_values(by='kid_id', ignore_index=True)
    combined.to_excel("kids_parents_from_kids_debug.xlsx", index=False)
    # print(f"Distinct parents from kids list: {kids_parents_from_kids['parent_name'].unique()}")

    # distinct_kids = kids_df['kid_name'].dropna().unique()
    # # put each parents name as key and value as list of kids names that contain the parents name, and it's class
    return combined

# all of them
def get_parent_kid_map(combined_df):
    parent_kid_map = {}

    # Step 1: Remove rows with empty or missing parent_name
    df_valid = combined_df[combined_df['parent_name'].notna() & (combined_df['parent_name'].str.strip() != '')]

    # Step 2: Group by parent_name and filter those with >=2 kids
    parent_groups = df_valid.groupby('parent_name').filter(lambda x: len(x) >= 2)

    # Step 3: Build the nested dictionary
    result = {}
    for parent, group in parent_groups.groupby('parent_name'):
        result[parent] = dict(zip(group['kid_name'], group['class']))

    # print(result)
    return result


data_map =  get_parent_kid_map(find_kids_of_parrents(parents_df, kids_df))

def getting_mount_from_string(amount_str):
    try:
        print(f"Extracting amount from string: {amount_str}")
        amount = int(''.join(filter(str.isdigit, amount_str)))
        return amount
    except:
        return 0.0

def calculate_months_paid(parents_df=parents_df):
    # Ensure 'Amount' is numeric
    parents_df['Amount'] = pd.to_numeric(parents_df['Amount'], errors='coerce')
    
    # Skip the first ROW (index 0)
    df_filtered = parents_df.iloc[1:].copy()
    
    parents_mount = dict(zip(df_filtered['parent_name'], df_filtered['Amount']))
    # print(parents_mount)
    print("Calculating months paid for each kid...")
    return parents_mount


amount_map = calculate_months_paid(parents_df)

print("Data maps prepared successfully.")
# print(f"Data Map (Parents to Kids): {data_map}")
# print(f"Amount Map (Parents to Amounts): {amount_map}")

# limiting row of kids_df
# infos
monthly_fee_per_kid_A = 25.0  # Example monthly fee per kid 25 for A5,.. and 15 for B0,...
monthly_fee_per_kid_B = 15.0  # Example monthly fee per kid 25 for A5,.. and 15 for B0,...
A5_names = ["A5","A6","A7","A8","A9","A10","A11","A12","G2"] # premair school, collige names 
B0_names = ["B0","B1","B2","B3","G1","G2"] # before primary school names , Sunday school

# all of the till here 

from openpyxl.styles import Color

def get_last_kid_update(sheet, df, row_idx):
    """Get the last update (month, text, color) for a single kid"""
    last_update = {"month": None, "text": None, "color": None}

    for month in reversed(months):
        col_idx = df.columns.get_loc(month) + 1
        cell = sheet.cell(row=row_idx, column=col_idx)
        text = str(cell.value).strip() if cell.value else ""

        color = None
        try:
            fill_color = cell.fill.start_color
            if isinstance(fill_color, Color):
                if fill_color.rgb and isinstance(fill_color.rgb, str):
                    color = fill_color.rgb
                elif fill_color.indexed is not None:
                    color = str(fill_color.indexed)
                elif fill_color.theme is not None:
                    color = f"theme:{fill_color.theme}"
                else:
                    color = None
            else:
                color = None
            
        except Exception:
            color = "FF595959"  # fallback safe color

        # Skip red cells (FFFF0000) and continue looking backwards
        if color == "FFFF0000":
            continue
        
        # Check if we have meaningful content
        if text or (color and color not in ["00000000", "None", ""]):
            if "Values must be of type <class 'int'>" in str(color):
                color = "FF595959"
            last_update = {"month": month, "text": text or None, "color": color}
            break

    return last_update



# 595959
def get_all_kids_last_updates(file_path):
    """Loop through all kids and get their last update"""
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    df = pd.read_excel(file_path)
    # remove the first row
    df = df.iloc[1:]
    # df = df.reset_index(drop=True)
    df.columns = ["kid_id", "kid_name", "parent_name", *months,
                  "class", "priceOn", "book_taken", "nabil_liste"]

    results = []

    for index, row in df.iterrows():
        kid_name = row['kid_name']
        if pd.isna(kid_name):
            continue  # skip empty rows

        # Convert DataFrame row index to Excel row number (+2 because header=None)
        excel_row_idx = index + 2
        update = get_last_kid_update(sheet, df, excel_row_idx)

        results.append({
            "kid_id": row["kid_id"],
            "kid_name": kid_name,
            "parent_name": row["parent_name"],
            "last_month": update["month"],
            "last_text": update["text"],
            "last_color": update["color"]
        })

    return pd.DataFrame(results)

kids_status = get_all_kids_last_updates("kids_list.xlsx")
# kids_status.to_excel("kids_last_updates_debug2.xlsx", index=False)


def get_monthly_fee_for_class(class_name):
    if class_name in A5_names:
        return monthly_fee_per_kid_A
    
    if class_name in B0_names:
        return monthly_fee_per_kid_B
    
    return monthly_fee_per_kid_A  # default to A fee



def determine_status_and_color(months_paid, monthly_fee, allocated_amount, class_name):
    if monthly_fee <= 0:
        return "Not yet registered", "FF595959"

    if allocated_amount == 0:
        return "Nothing paid.", "FFFF0000"
    
    if months_paid >= 1.0:
        return "Fully paid.", "FF92D050"
    
    if class_name.strip() == 'A5' and abs(allocated_amount - 15) < 0.01:
        return "G1 and G2 paid €15 instead of €25.", "FFFFFF00"

    if class_name.strip() == 'A5':
        if allocated_amount in [10, 15, 20]:
            return "Transfers only €10, €15, or €20 instead of €25.", "FFC65911"
    elif class_name.strip() == 'B0':
        if allocated_amount in [10]:
            return "Transfers only €10 instead of €15.", "FFC65911"

    return f"Partial payment: {allocated_amount:.2f}€ ({months_paid:.2f} months)", "FFFFC000"

# Normalize: store colors in uppercase for consistency (openpyxl uses uppercase,
# but your code has lowercase — let's accept both)
_STATUS_COLOR_MAP = [
    ("Not yet registered", "FF595959"),
    ("Nothing paid.", "FFFF0000"),
    ("Fully paid.", "FF92D050"),
    ("G1 and G2 paid €15 instead of €25.", "FFFFFF00"),
    ("Transfers only €10, €15, or €20 instead of €25.", "FFC65911"),
    ("Transfers only €10 instead of €15.", "FFC65911"),
]

# Build lookup dicts (handle case-insensitive color matching)
TEXT_TO_COLOR = {text: color.upper() for text, color in _STATUS_COLOR_MAP}
COLOR_TO_TEXT = {color.upper(): text for text, color in _STATUS_COLOR_MAP}

# Special color for partial payments
PARTIAL_COLOR = "FFFFC000"
PARTIAL_COLOR_UPPER = PARTIAL_COLOR.upper()
GENERIC_PARTIAL_TEXT = "Partial payment"

def text_to_color(status_text: str) -> str:
    """Map status text to ARGB color string (8-digit, uppercase)."""
    # Check exact matches
    if status_text in TEXT_TO_COLOR:
        return TEXT_TO_COLOR[status_text]
    
    # Check if it's a partial payment (starts with known prefix)
    if status_text.startswith("Partial payment:"):
        return PARTIAL_COLOR_UPPER
    
    # Fallback: unknown → dark gray
    return "FF595959"


def color_to_text(color: str) -> str:
    """Map ARGB color string (6 or 8 digit, any case) to status text."""
    if not color:
        return "Unknown"
    
    color = color.strip().upper()
    
    # If 6-digit, assume opaque (prepend FF)
    if len(color) == 6:
        color = "FF" + color
    elif len(color) != 8:
        return "Unknown"
    
    # Exact match?
    if color in COLOR_TO_TEXT:
        return COLOR_TO_TEXT[color]
    
    # Partial payment color?
    if color == PARTIAL_COLOR_UPPER:
        return GENERIC_PARTIAL_TEXT
    
    # Fallback
    return "Unknown"


'''
#c65911 :Transfers only €10, €15, or €20 instead of €25.
#ff0000 :Nothing paid.
#92d050 :Fully paid.
#ffff00 : G1 and G2 paid €15 instead of €25.
#f4b084 :Unenrolled kids.
'''

def calculate_kid_payments(data_map, amount_map, kid_status):
    """
    kid_status: dict {kid_name: {'allocated_amount': float, 'class': str, 'monthly_fee': float}}
    """
    kid_payment_status = {}

    # Build prior allocation map per kid and per parent
    prior_kid_alloc = {}
    prior_parent_total = {}
    for kid_name, info in kid_status.items():
        alloc = float(info.get('allocated_amount', 0.0))
        prior_kid_alloc[kid_name] = alloc
        parent = info.get('parent', '')
        if parent:
            prior_parent_total[parent] = prior_parent_total.get(parent, 0.0) + alloc

    for parent, kids in data_map.items():
        new_payment = float(amount_map.get(parent, 0.0))
        prior_total = prior_parent_total.get(parent, 0.0)
        total_effective_amount = prior_total + new_payment

        print(f"\nParent: {parent}, New Payment: €{new_payment}, Prior Total: €{prior_total}, Effective Total: €{total_effective_amount}")

        # Get monthly fees
        kid_list = list(kids.items())
        kid_fees = {}
        total_monthly_fee = 0.0
        for kid, cls in kid_list:
            fee = get_monthly_fee_for_class(cls)
            kid_fees[kid] = fee
            total_monthly_fee += fee

        print(f"Total Monthly Fee for {parent}: €{total_monthly_fee}")

        if total_monthly_fee <= 0:
            for kid_name, class_name in kid_list:
                status_msg, color = determine_status_and_color(0, 0, 0, class_name)
                kid_payment_status[kid_name] = {
                    'parent': parent,
                    'class': class_name.strip(),
                    'monthly_fee': 0.0,
                    'allocated_amount': 0.0,
                    'months_paid': 0.0,
                    'status': status_msg,
                    'color': color,
                    'extras': 0.0,
                    'extras_color': color
                }
            continue

        # Compute full months and remainder from TOTAL effective amount
        full_months_total = int(total_effective_amount // total_monthly_fee)
        remainder = total_effective_amount - (full_months_total * total_monthly_fee)

        # Base allocation per kid
        base_allocations = {
            kid_name: full_months_total * kid_fees[kid_name]
            for kid_name, _ in kid_list
        }

        # Assign remainder to FIRST kid (as before)
        allocations = {}
        for i, (kid_name, _) in enumerate(kid_list):
            if i == 0:
                allocations[kid_name] = base_allocations[kid_name] + remainder
            else:
                allocations[kid_name] = base_allocations[kid_name]

        # Now build result using CUMULATIVE allocations
        for kid_name, class_name in kid_list:
            monthly_fee = kid_fees[kid_name]
            allocated = allocations[kid_name]  # cumulative

            months_paid = allocated / monthly_fee if monthly_fee > 0 else 0.0

            status_msg, color = determine_status_and_color(
                months_paid, monthly_fee, allocated, class_name
            )

            if monthly_fee > 0:
                full_months_for_kid = int(allocated // monthly_fee)
                extras = allocated - (full_months_for_kid * monthly_fee)
                extras = round(extras, 2)
                extras_color = "#ffc000" if extras > 1e-2 else "#92d050"
            else:
                extras = 0.0
                extras_color = color

            allocated = round(allocated, 2)
            months_paid = round(months_paid, 2)
            monthly_fee = round(monthly_fee, 2)

            kid_payment_status[kid_name] = {
                'parent': parent,
                'class': class_name.strip(),
                'monthly_fee': monthly_fee,
                'allocated_amount': allocated,          # cumulative
                'months_paid': months_paid,
                'status': status_msg,
                'color': color,
                'extras': extras,
                'extras_color': extras_color
            }

            print(f"  → {kid_name}: €{allocated:.2f} allocated → {months_paid:.2f} months → {status_msg}")
            if extras > 0:
                print(f"      (Extras: €{extras:.2f})")

    return kid_payment_status

print("\nCalculating kid payment statuses...-------------------------")
# calculate_kid_payments(data_map, amount_map , kids_status )

def copy_cell_format(source_cell, target_cell):
    """Copy all formatting from source cell to target cell"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def extend_months_to_2years():
    """Extend the months list to cover 2 years instead of 1.5 years"""
    months_2years = [
        '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8',  # Year 1
        '9_next', '10_next', '11_next', '12_next', '1_next', '2_next', 
        '3_next', '4_next', '5_next', '6_next', '7_next', '8_next'  # Year 2
    ]
    return months_2years

def update_excel_with_payments(kids_df, kid_payment_status, kids_first_rows, kids_last_rows, 
                                output_file="kids_list_updated.xlsx"):
    """
    Update the Excel file with new payment statuses while preserving formatting
    """
    # Load the original workbook to preserve formatting
    
    
    # Extend months to 2 years
    
    months_extended = extend_months_to_2years()

    # Calculate how many new month columns to add
    original_month_count = len(months)
    new_month_count = len(months_extended)
    months_to_add = new_month_count - original_month_count
    
    # Get the position where month columns start (after kid_id, kid_name, parent_name)
    month_start_col = 4  # Column D (1-indexed)
    month_end_col = month_start_col + original_month_count - 1
    
    # Get a reference cell from existing months to copy style from
    reference_month_cell = ws.cell(row=4, column=month_start_col)

    # Insert new columns after the existing month columns
    if months_to_add > 0:
        for _ in range(months_to_add):
            ws.insert_cols(month_end_col + 1)
        
        # Find where 2026 starts (after the 24 months, which is 9_next to 8_next)
        year_2026_start_col = month_start_col + 16   # After 24 months
        
        # Merge cells for "2026" header (row 1)
        ws.merge_cells(start_row=2, start_column=year_2026_start_col, 
                    end_row=2, end_column=year_2026_start_col + months_to_add )
        
        # Set "2026" text and copy style from "2025" cell
        year_2026_cell = ws.cell(row=2, column=year_2026_start_col)
        year_2026_cell.value = "2026"
        
        # Copy format from "2025" header
        year_2025_cell = ws.cell(row=2, column=month_start_col + 4)  # Where 2025 starts
        # copy_cell_format(year_2025_cell, year_2026_cell)
        # outline the merged cell
        bold_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        year_2026_cell.border = bold_border
        
        # Update headers for new month columns (row 2 and row 3)
        for i, month in enumerate(months_extended):
            col_idx = month_start_col + i
            
            # Row 2: Month numbers (9, 10, 11, 12, 1, 2, etc.)
            month_num_cell = ws.cell(row=3, column=col_idx)
            
            # Row 3: Month values (25, 25, 25, etc.)
            header_cell = ws.cell(row=3, column=col_idx)
            
            # For new columns (beyond original), copy format and set values
            if i >= original_month_count:
                # Get the month number (remove "_next" suffix for display)
                display_month = month.replace('_next', '') 
                print(f"Setting new month column {col_idx} for month '{month}' as '{display_month}'")
                month_num_cell.value = display_month
                header_cell.value = display_month # was month need fix
                
                # Copy format from corresponding month in year 1
                # e.g., 9_next copies from 9, 10_next from 10, etc.
                source_col_offset = i % 12  # Cycle through 12 months
                source_col = month_start_col + source_col_offset
                
                source_month_num = ws.cell(row=2, column=source_col)
                source_header = ws.cell(row=3, column=source_col)
                
                copy_cell_format(source_month_num, month_num_cell)
                copy_cell_format(source_header, header_cell)
    

    # Get kids status info
    kids_status_dict = {}
    for _, row in kids_status.iterrows():
        kid_name = row['kid_name']
        kids_status_dict[kid_name] = {
            'last_month': row['last_month'],
            'last_text': row['last_text'],
            'last_color': row['last_color']
        }
    

    # Update parent names and add phone numbers
    # Find the parent_name column (should be column 3)
    parent_name_col = 3
    
    # Get the combined kids_parents data with phone numbers
    kids_parents_combined = find_kids_of_parrents(parents_df, kids_df)
    
    # Create a mapping of kid_id to parent info
    kid_to_parent_info = {}
    for _, row in kids_parents_combined.iterrows():
        kid_to_parent_info[row['kid_id']] = {
            'parent_name': row['parent_name'],
            'phone_number': row.get('phone_number', '')
        }
    
    # Insert a new column for phone numbers after parent_name (before months)
    # This will be column 4, and months will shift to column 5
    last_colomn = ws.max_column
    ws.insert_cols(last_colomn + 1)
    
    # Update the header for phone number column
    phone_header_cell = ws.cell(row=3, column=last_colomn + 1)
    phone_header_cell.value = "Phone Number"
    # Copy format from parent_name header
    parent_header_cell = ws.cell(row=3, column=parent_name_col)
    copy_cell_format(parent_header_cell, phone_header_cell)
    
    # Also update row 1 and row 2 for the phone number column
    # phone_row1_cell = ws.cell(row=3, column=last_colomn + 1)
    phone_row2_cell = ws.cell(row=3, column=last_colomn + 1)
    # copy_cell_format(ws.cell(row=3, column=parent_name_col), phone_row1_cell)
    copy_cell_format(ws.cell(row=2, column=parent_name_col), phone_row2_cell)

    # Process each kid row (starting from row 4)
    start_row = 4
    for idx, kid_row in kids_df.iterrows():
        excel_row = start_row + idx
        kid_name = kid_row['kid_name']
        kid_id = kid_row['kid_id']

        if pd.isna(kid_name):
            continue

        # Update parent name and add phone number
        if kid_id in kid_to_parent_info:
            parent_info = kid_to_parent_info[kid_id]
            
            # Update parent name (column 3)
            parent_cell = ws.cell(row=excel_row, column=parent_name_col)
            parent_cell.value = parent_info['parent_name'] if parent_info['parent_name'] != 'nan' else parent_cell.value
            
            # Add phone number (column 4)
            phone_cell = ws.cell(row=excel_row, column=last_colomn + 1)
            phone_cell.value = parent_info['phone_number']
            # Copy format from parent cell
            copy_cell_format(parent_cell, phone_cell)

        # Get payment status for this kid
        payment_info = kid_payment_status.get(kid_name, {})
        
        if not payment_info:
            print(f"⚠️ No payment info found for {kid_name}, skipping updates")
            continue
        
        # Get kid's last status
        last_status = kids_status_dict.get(kid_name, {})
        last_month = last_status.get('last_month')
        last_color = last_status.get('last_color')
        last_text = last_status.get('last_text')
        # After retrieving last_status:
        last_color = (last_status.get('last_color') or "").upper().replace("#", "")
        is_not_registered = last_color in ["FF595959", "595959"]

        # If kid is not registered and no payment, leave all future months untouched
        if is_not_registered and (payment_info.get('allocated_amount', 0) == 0):
            print(f"⚠️ Skipping updates for unregistered kid with no payment: {kid_name}")
            continue  # Skip updating this row entirely

        # Find the index of the last updated month
        try:
            last_month_idx = months_extended.index(last_month) if last_month else -1
        except ValueError:
            last_month_idx = -1
        
        # Determine how many new months to fill based on payment
        months_paid = payment_info.get('months_paid', 0.0)
        full_months_paid = int(months_paid)
        monthly_fee = payment_info.get('monthly_fee', 0.0)
        
        # Determine the color for new entries
        new_color = payment_info.get('color', 'FF595959')
        new_status = payment_info.get('status', '')
        
        # Calculate extras
        extras = payment_info.get('extras', 0.0)
        extras_color = payment_info.get('extras_color', '#92d050')
        
        # Fill month columns
        for i, month in enumerate(months_extended):
            col_idx = month_start_col + i
            cell = ws.cell(row=excel_row, column=col_idx)
            
            # Keep old data as-is (before or at last_month_idx)
            if i <= last_month_idx:
                # Preserve existing formatting and value
                continue
            
            # Fill new months based on payment
            elif i <= last_month_idx + full_months_paid:
                # Put the monthly fee price
                if monthly_fee > 0:
                    cell.value = int(monthly_fee) if monthly_fee == int(monthly_fee) else monthly_fee
                else:
                    cell.value = ""
                
                # Copy style from reference cell
                copy_cell_format(reference_month_cell, cell)
                
                # Apply payment status color
                cell.fill = PatternFill(start_color=new_color.replace("#", ""), 
                                       end_color=new_color.replace("#", ""), 
                                       fill_type="solid")
            
            # Handle extras in the next month
            elif i == last_month_idx + full_months_paid + 1 and extras > 0:
                cell.value = extras if extras != int(extras) else int(extras)
                
                # Copy style from reference cell
                copy_cell_format(reference_month_cell, cell)
                
                # Apply extras color
                cell.fill = PatternFill(start_color=extras_color.replace("#", ""), 
                                       end_color=extras_color.replace("#", ""), 
                                       fill_type="solid")
            
            else:
                # Mark ONLY the first unpaid month in red, then stop
                if i == last_month_idx + full_months_paid + 1 and extras == 0:
                    cell.fill = PatternFill(start_color="FFFF0000",
                                        end_color="FFFF0000",
                                        fill_type="solid")
        
        print(f"✅ Updated {kid_name}: {full_months_paid} months paid (€{monthly_fee}/month), extras: €{extras}")
    
    # Save the updated workbook
    wb.save(output_file)
    print(f"\n✅ Excel file updated successfully: {output_file}")
    return output_file

# Add this at the end of your script, after calculate_kid_payments
print("\n" + "="*60)
print("UPDATING EXCEL FILE WITH NEW PAYMENT STATUS")
print("="*60 + "\n")

# Calculate kid payment statuses
kid_payment_status = calculate_kid_payments(data_map, amount_map, 
                                            {row['kid_name']: {
                                                'allocated_amount': 0.0,  # Start fresh or use previous
                                                'class': row['class'],
                                                'monthly_fee': get_monthly_fee_for_class(row['class']),
                                                'parent': row['parent_name']
                                            } for _, row in kids_df.iterrows()})

# Update the Excel file
output_file = update_excel_with_payments(
    kids_df=kids_df,
    kid_payment_status=kid_payment_status,
    kids_first_rows=kids_first_rows,
    kids_last_rows=kids_last_rows,
    output_file="kids_list_updated.xlsx"
)

print(f"\n🎉 Process completed! Check '{output_file}' for results.")