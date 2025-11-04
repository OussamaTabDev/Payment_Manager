import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side
from openpyxl import load_workbook
import pandas as pd
from copy import copy

# ============================================================================
# CONSTANTS
# ============================================================================

# File paths
PARENT_FILE = "parents_payments.xlsx"
KID_FILE = "theone.xlsx"
OUTPUT_FILE = "kids_list_updated3.xlsx"

# Monthly fees
MONTHLY_FEE_A = 25.0  # Primary school, college
MONTHLY_FEE_B = 15.0  # Before primary school, Sunday school

# Class names
A5_NAMES = ["A5", "A6", "A7", "A8", "A9", "A10", "A11", "A12", "G2"]
B0_NAMES = ["B0", "B1", "B2", "B3", "G1", "G2"]

# Months mapping
MONTHS_1_5_YEARS = [
    '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8', '9_next',
    '10_next', '11_next', '12_next', '1_next'
]

MONTHS_2_YEARS = [
    '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8',
    '9_next', '10_next', '11_next', '12_next', '1_next', '2_next',
    '3_next', '4_next', '5_next', '6_next', '7_next', '8_next'
]

# Status color mapping
STATUS_COLOR_MAP = [
    ("Not yet registered", "FF595959"),
    ("Nothing paid.", "FFFF0000"),
    ("Fully paid.", "FF92D050"),
    ("G1 and G2 paid €15 instead of €25.", "FFFFFF00"),
    ("Transfers only €10, €15, or €20 instead of €25.", "FFC65911"),
    ("Transfers only €10 instead of €15.", "FFC65911"),
]

TEXT_TO_COLOR = {text: color.upper() for text, color in STATUS_COLOR_MAP}
COLOR_TO_TEXT = {color.upper(): text for text, color in STATUS_COLOR_MAP}
PARTIAL_COLOR = "FFFFC000"

# Mode setting
MODE = "prod"  # Change to "test" for testing with limited rows

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def load_data(parent_file, kid_file):
    """Load parent and kid data from Excel files."""
    parents_df = pd.read_excel(parent_file, header=None)
    kids_df = pd.read_excel(kid_file, header=None)
    
    kids_first_rows = kids_df.iloc[:3]
    kids_df = kids_df.iloc[3:]
    print(kids_df.head())
    kids_df = kids_df.reset_index(drop=True)
    print(f"Kids before filtering:")
    #exporting 
    kids_df.to_excel("kids_debug_before_filtering.xlsx", index=False)
    print("---------")
    parents_df.columns = [
        "Account_Number", "Booking_Date", "Value_Date", "Transaction_Text",
        "Usage_Purpose", "parent_name", "Account_or_IBAN", "BIC_SWIFT_Code",
        "Amount", "Currency", "Info"
    ]
    
    wb = load_workbook(kid_file)
    ws = wb.active
    last_column = ws.max_column
    # print(f"Last column in kids file: {last_column}")
    months = MONTHS_1_5_YEARS if last_column < 25 else MONTHS_2_YEARS
    
    if last_column < 25:
        kids_df.columns = [
            "kid_id", 'kid_name', 'parent_name', *months,
            'class', 'priceOn', 'book_taken', 'nabil_liste'
        ]
    else:
        kids_df.columns = [
            "kid_id", 'kid_name', 'parent_name', *months,
            'class', 'priceOn', 'book_taken', 'nabil_liste', 'phone_number'
        ]
    
    return parents_df, kids_df, kids_first_rows, months


def filter_dataframe(kids_df, mode):
    """Filter DataFrame based on mode (test or prod)."""
    backup_kids_df = kids_df.copy()
    kids_last_rows = kids_df
    
    if mode == "test":
        kids_df = kids_df.head(100)
        print(f"🧪 Testing mode: Limited to {len(kids_df)} rows.")
    else:
        if "kid_id" in kids_df.columns:
            stop_index = kids_df["kid_id"].isna().idxmax() if kids_df["kid_id"].isna().any() else None
            
            if stop_index is not None and stop_index > 0:
                kids_last_rows = kids_df.iloc[stop_index:]
                kids_df = kids_df.iloc[:stop_index]
                print(f"✅ Production mode: Stopped at first empty kid_id (row {stop_index}).")
            else:
                print("✅ Production mode: No missing kid_id found. Using all rows.")
    
    return kids_df, kids_last_rows, backup_kids_df


def get_monthly_fee_for_class(class_name):
    """Get monthly fee based on class name."""
    if class_name in A5_NAMES:
        return MONTHLY_FEE_A
    if class_name in B0_NAMES:
        return MONTHLY_FEE_B
    return MONTHLY_FEE_A


def text_to_color(status_text):
    """Map status text to ARGB color string."""
    if status_text in TEXT_TO_COLOR:
        return TEXT_TO_COLOR[status_text]
    if status_text.startswith("Partial payment:"):
        return PARTIAL_COLOR
    return "FF595959"


def color_to_text(color):
    """Map ARGB color string to status text."""
    if not color:
        return "Unknown"
    
    color = color.strip().upper()
    if len(color) == 6:
        color = "FF" + color
    elif len(color) != 8:
        return "Unknown"
    
    if color in COLOR_TO_TEXT:
        return COLOR_TO_TEXT[color]
    if color == PARTIAL_COLOR:
        return "Partial payment"
    return "Unknown"


def copy_cell_format(source_cell, target_cell):
    """Copy all formatting from source cell to target cell."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


# ============================================================================
# CORE PROCESSING FUNCTIONS
# ============================================================================

def find_kids_of_parents(parents_df, kids_df, backup_kids_df):
    """Find and match kids with their parents."""
    distinct_parents = parents_df['parent_name'].dropna().unique()
    kids_parents_from_kids = kids_df[['kid_id', 'kid_name', 'parent_name', 'class']].copy()
    
    kids_parents_from_kids['phone_number'] = kids_parents_from_kids['parent_name'].str.extract(r'\(([^)]*)\)')
    kids_parents_from_kids['parent_name'] = kids_parents_from_kids['parent_name'].str.replace(r'\s*\([^\)]*\)', '', regex=True)
    kids_parents_from_kids['parent_name'] = kids_parents_from_kids['parent_name'].str.strip()
    
    empty_kids_parents = kids_parents_from_kids[kids_parents_from_kids['parent_name'].isna()]
    kids_parents_from_kids = kids_parents_from_kids[kids_parents_from_kids['parent_name'].notna()]
    other_distinct_parents = kids_parents_from_kids['parent_name'].dropna().unique()
    
    # Complete missing parent names
    for index, row in empty_kids_parents.iterrows():
        kid_name = row['kid_name']
        matched_parent = None
        
        for parent in other_distinct_parents:
            if not isinstance(parent, str) or not parent.strip():
                continue
            last_name_parent = parent.split()[-1]
            if kid_name and kid_name.split() and last_name_parent.lower() == kid_name.split()[0].lower():
                matched_parent = parent
                break
        
        if not matched_parent:
            for parent in distinct_parents:
                if not isinstance(parent, str) or not parent.strip():
                    continue
                last_name_parent = parent.split()[-1]
                if kid_name and kid_name.split() and last_name_parent.lower() == kid_name.split()[0].lower():
                    matched_parent = parent
                    break
        
        if matched_parent:
            empty_kids_parents.at[index, 'parent_name'] = matched_parent
        else:
            for kid_id in empty_kids_parents['kid_id']:
                idx = empty_kids_parents.index[empty_kids_parents['kid_id'] == kid_id][0]
                if pd.isna(empty_kids_parents.at[idx, 'parent_name']) or empty_kids_parents.at[idx, 'parent_name'] == '':
                    parent_name = backup_kids_df.loc[backup_kids_df['kid_id'] == kid_id, 'parent_name'].values
                    if len(parent_name) > 0:
                        empty_kids_parents.at[idx, 'parent_name'] = str(parent_name[0])
    
    # Replace parent names with matching distinct parents
    for index, row in kids_parents_from_kids.iterrows():
        current_parent_name = row['parent_name']
        matched_parent = None
        
        for parent in distinct_parents:
            try:
                parent_str = str(parent)
                current_parent_name_str = str(current_parent_name)
                if parent_str in current_parent_name_str or current_parent_name_str in parent_str and len(parent_str) > 3:
                    matched_parent = parent
                    break
            except TypeError:
                continue
        
        if matched_parent:
            kids_parents_from_kids.at[index, 'parent_name'] = matched_parent
    
    combined = pd.concat([kids_parents_from_kids, empty_kids_parents], ignore_index=True)
    combined = combined.sort_values(by='kid_id', ignore_index=True)
    
    return combined


def get_parent_kid_map(combined_df):
    """Create mapping of parents to their kids."""
    df_valid = combined_df[combined_df['parent_name'].notna() & (combined_df['parent_name'].str.strip() != '')]
    parent_groups = df_valid.groupby('parent_name').filter(lambda x: len(x) >= 2)
    print(f"Found {parent_groups['parent_name'].nunique()} parents with 2 or more kids.")
    print(df_valid.head())
    result = {}
    for parent, group in parent_groups.groupby('parent_name'):
        result[parent] = dict(zip(group['kid_name'], group['class']))
    
    return result


def calculate_months_paid(parents_df):
    """Calculate total amount paid by each parent."""
    parents_df['Amount'] = pd.to_numeric(parents_df['Amount'], errors='coerce')
    df_filtered = parents_df.iloc[1:].copy()
    parents_amount = dict(zip(df_filtered['parent_name'], df_filtered['Amount']))
    return parents_amount


def get_last_kid_update(sheet, df, row_idx, months):
    """Get the last update (month, text, color) for a single kid."""
    last_update = {"month": None, "text": None, "color": None}
    
    for month in reversed(months):
        col_idx = df.columns.get_loc(month) + 1
        cell = sheet.cell(row=row_idx, column=col_idx)
        text = str(cell.value).strip() if cell.value else ""
        
        color = None
        try:
            fill_color = cell.fill.start_color
            if isinstance(fill_color, Color):
                if fill_color.rgb and isinstance(fill_color.rgb, str):
                    color = fill_color.rgb
                elif fill_color.indexed is not None:
                    color = str(fill_color.indexed)
                elif fill_color.theme is not None:
                    color = f"theme:{fill_color.theme}"
            else:
                color = None
        except Exception:
            color = "FF595959"
        
        if color == "FFFF0000":
            continue
        
        if text or (color and color not in ["00000000", "None", ""]):
            if "Values must be of type <class 'int'>" in str(color):
                color = "FF595959"
            last_update = {"month": month, "text": text or None, "color": color}
            break
    
    return last_update


def get_all_kids_last_updates(file_path, months):
    """Get last update for all kids."""
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    df = pd.read_excel(file_path)
    last_column = sheet.max_column
    df = df.iloc[1:]
    
    if last_column < 25:
        df.columns = ["kid_id", 'kid_name', 'parent_name', *months, 'class', 'priceOn', 'book_taken', 'nabil_liste']
    else:
        df.columns = ["kid_id", 'kid_name', 'parent_name', *months, 'class', 'priceOn', 'book_taken', 'nabil_liste', 'phone_number']
    
    results = []
    for index, row in df.iterrows():
        kid_name = row['kid_name']
        if pd.isna(kid_name):
            continue
        
        excel_row_idx = index + 2
        update = get_last_kid_update(sheet, df, excel_row_idx, months)
        
        results.append({
            "kid_id": row["kid_id"],
            "kid_name": kid_name,
            "parent_name": row["parent_name"],
            "last_month": update["month"],
            "last_text": update["text"],
            "last_color": update["color"]
        })
    
    return pd.DataFrame(results)


def determine_status_and_color(months_paid, monthly_fee, allocated_amount, class_name):
    """Determine payment status and color based on amount paid."""
    if monthly_fee <= 0:
        return "Not yet registered", "FF595959"
    
    if allocated_amount == 0:
        return "Nothing paid.", "FFFF0000"
    
    if months_paid >= 1.0:
        return "Fully paid.", "FF92D050"
    
    if class_name.strip() == 'A5' and abs(allocated_amount - 15) < 0.01:
        return "G1 and G2 paid €15 instead of €25.", "FFFFFF00"
    
    if class_name.strip() == 'A5':
        if allocated_amount in [10, 15, 20]:
            return "Transfers only €10, €15, or €20 instead of €25.", "FFC65911"
    elif class_name.strip() == 'B0':
        if allocated_amount in [10]:
            return "Transfers only €10 instead of €15.", "FFC65911"
    
    return f"Partial payment: {allocated_amount:.2f}€ ({months_paid:.2f} months)", "FFFFC000"


def calculate_kid_payments(data_map, amount_map, kid_status):
    """
    kid_status: dict {kid_name: {'allocated_amount': float, 'class': str, 'monthly_fee': float}}
    """
    kid_payment_status = {}

    # Build prior allocation map per kid and per parent
    prior_kid_alloc = {}
    prior_parent_total = {}
    for kid_name, info in kid_status.items():
        alloc = float(info.get('allocated_amount', 0.0))
        prior_kid_alloc[kid_name] = alloc
        parent = info.get('parent', '')
        if parent:
            prior_parent_total[parent] = prior_parent_total.get(parent, 0.0) + alloc
    print(data_map)
    for parent, kids in data_map.items():
        new_payment = float(amount_map.get(parent, 0.0))
        prior_total = prior_parent_total.get(parent, 0.0)
        total_effective_amount = prior_total + new_payment

        # print(f"\nParent: {parent}, New Payment: €{new_payment}, Prior Total: €{prior_total}, Effective Total: €{total_effective_amount}")

        # Get monthly fees
        kid_list = list(kids.items())
        kid_fees = {}
        total_monthly_fee = 0.0
        for kid, cls in kid_list:
            fee = get_monthly_fee_for_class(cls)
            kid_fees[kid] = fee
            total_monthly_fee += fee

        # print(f"Total Monthly Fee for {parent}: €{total_monthly_fee}")

        if total_monthly_fee <= 0:
            for kid_name, class_name in kid_list:
                status_msg, color = determine_status_and_color(0, 0, 0, class_name)
                kid_payment_status[kid_name] = {
                    'parent': parent,
                    'class': class_name.strip(),
                    'monthly_fee': 0.0,
                    'allocated_amount': 0.0,
                    'months_paid': 0.0,
                    'status': status_msg,
                    'color': color,
                    'extras': 0.0,
                    'extras_color': color
                }
            continue

        # Compute full months and remainder from TOTAL effective amount
        full_months_total = int(total_effective_amount // total_monthly_fee)
        remainder = total_effective_amount - (full_months_total * total_monthly_fee)

        # Base allocation per kid
        base_allocations = {
            kid_name: full_months_total * kid_fees[kid_name]
            for kid_name, _ in kid_list
        }

        # Assign remainder to FIRST kid (as before)
        allocations = {}
        for i, (kid_name, _) in enumerate(kid_list):
            if i == 0:
                allocations[kid_name] = base_allocations[kid_name] + remainder
            else:
                allocations[kid_name] = base_allocations[kid_name]

        # Now build result using CUMULATIVE allocations
        for kid_name, class_name in kid_list:
            monthly_fee = kid_fees[kid_name]
            allocated = allocations[kid_name]  # cumulative

            months_paid = allocated / monthly_fee if monthly_fee > 0 else 0.0

            status_msg, color = determine_status_and_color(
                months_paid, monthly_fee, allocated, class_name
            )

            if monthly_fee > 0:
                full_months_for_kid = int(allocated // monthly_fee)
                extras = allocated - (full_months_for_kid * monthly_fee)
                extras = round(extras, 2)
                extras_color = "#ffc000" if extras > 1e-2 else "#92d050"
            else:
                extras = 0.0
                extras_color = color

            allocated = round(allocated, 2)
            months_paid = round(months_paid, 2)
            monthly_fee = round(monthly_fee, 2)

            kid_payment_status[kid_name] = {
                'parent': parent,
                'class': class_name.strip(),
                'monthly_fee': monthly_fee,
                'allocated_amount': allocated,          # cumulative
                'months_paid': months_paid,
                'status': status_msg,
                'color': color,
                'extras': extras,
                'extras_color': extras_color
            }

            # print(f"  → {kid_name}: €{allocated:.2f} allocated → {months_paid:.2f} months → {status_msg}")
            # if extras > 0:
            #     print(f"      (Extras: €{extras:.2f})")

    return kid_payment_status


def update_excel_with_payments(kids_df, kid_payment_status, kids_status, months, kid_file, output_file):
    """Update Excel file with payment statuses."""
    wb = load_workbook(kid_file)
    ws = wb.active
    
    months_extended = MONTHS_2_YEARS
    original_month_count = len(months)
    new_month_count = len(months_extended)
    months_to_add = new_month_count - original_month_count
    
    month_start_col = 4
    month_end_col = month_start_col + original_month_count - 1
    reference_month_cell = ws.cell(row=4, column=month_start_col)
    
    last_expected_col = month_start_col + new_month_count - 1
    already_extended = False
    
    if ws.max_column >= last_expected_col:
        last_month_header = ws.cell(row=3, column=last_expected_col).value
        if last_month_header and str(last_month_header).strip() in ['8', '8_next']:
            already_extended = True
            print("✅ Months already extended to 2 years.")
    
    if months_to_add > 0 and not already_extended:
        print(f"➕ Adding {months_to_add} new month columns...")
        for _ in range(months_to_add):
            ws.insert_cols(month_end_col + 1)
        
        year_2026_start_col = month_start_col + 16
        ws.merge_cells(start_row=2, start_column=year_2026_start_col, end_row=2, end_column=year_2026_start_col + months_to_add)
        year_2026_cell = ws.cell(row=2, column=year_2026_start_col)
        
        bold_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        year_2026_cell.border = bold_border
        
        for i, month in enumerate(months_extended):
            col_idx = month_start_col + i
            month_num_cell = ws.cell(row=3, column=col_idx)
            
            if i >= original_month_count:
                display_month = month.replace('_next', '')
                month_num_cell.value = display_month
                source_col_offset = i % 12
                source_col = month_start_col + source_col_offset
                source_month_num = ws.cell(row=3, column=source_col)
                copy_cell_format(source_month_num, month_num_cell)
    
    kids_status_dict = {}
    for _, row in kids_status.iterrows():
        kid_name = row['kid_name']
        kids_status_dict[kid_name] = {
            'last_month': row['last_month'],
            'last_text': row['last_text'],
            'last_color': row['last_color']
        }
    
    parent_name_col = 3
    last_column = ws.max_column
    phone_col_exists = ws.max_column >= 32
    phone_number_column = last_column
    
    if not phone_col_exists:
        print("➕ Adding phone number column...")
        ws.insert_cols(last_column + 1)
        phone_header_cell = ws.cell(row=3, column=phone_number_column + 1)
        phone_header_cell.value = "Phone Number"
        parent_header_cell = ws.cell(row=3, column=parent_name_col)
        copy_cell_format(parent_header_cell, phone_header_cell)
    
    start_row = 4
    for idx, kid_row in kids_df.iterrows():
        excel_row = start_row + idx
        kid_name = kid_row['kid_name']
        
        if pd.isna(kid_name):
            continue
        
        payment_info = kid_payment_status.get(kid_name, {})
        if not payment_info:
            continue
        
        last_status = kids_status_dict.get(kid_name, {})
        last_month = last_status.get('last_month')
        last_color = (last_status.get('last_color') or "").upper().replace("#", "")
        is_not_registered = last_color in ["FF595959", "595959"]
        
        if is_not_registered and (payment_info.get('allocated_amount', 0) == 0):
            continue
        
        try:
            last_month_idx = months_extended.index(last_month) if last_month else -1
        except ValueError:
            last_month_idx = -1
        
        months_paid = payment_info.get('months_paid', 0.0)
        full_months_paid = int(months_paid)
        monthly_fee = payment_info.get('monthly_fee', 0.0)
        new_color = payment_info.get('color', 'FF595959')
        extras = payment_info.get('extras', 0.0)
        extras_color = payment_info.get('extras_color', '#92d050')
        
        for i, month in enumerate(months_extended):
            col_idx = month_start_col + i
            cell = ws.cell(row=excel_row, column=col_idx)
            
            if i <= last_month_idx:
                continue
            elif i <= last_month_idx + full_months_paid:
                if monthly_fee > 0:
                    cell.value = int(monthly_fee) if monthly_fee == int(monthly_fee) else monthly_fee
                else:
                    cell.value = ""
                copy_cell_format(reference_month_cell, cell)
                cell.fill = PatternFill(start_color=new_color.replace("#", ""), end_color=new_color.replace("#", ""), fill_type="solid")
            elif i == last_month_idx + full_months_paid + 1 and extras > 0:
                cell.value = extras if extras != int(extras) else int(extras)
                copy_cell_format(reference_month_cell, cell)
                cell.fill = PatternFill(start_color=extras_color.replace("#", ""), end_color=extras_color.replace("#", ""), fill_type="solid")
            else:
                if i == last_month_idx + full_months_paid + 1 and extras == 0:
                    cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    wb.save(output_file)
    print(f"\n✅ Excel file updated successfully: {output_file}")
    return output_file


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function."""
    print("="*60)
    print("PAYMENT PROCESSING SYSTEM")
    print("="*60 + "\n")
    
    # Load data
    print("📂 Loading data...")
    parents_df, kids_df, kids_first_rows, months = load_data(PARENT_FILE, KID_FILE)
    print("✅ Data loaded successfully.\n")
    print("before:" ,kids_df.head())
    # Filter DataFrame
    kids_df, kids_last_rows, backup_kids_df = filter_dataframe(kids_df, MODE)
    print(kids_df.head())
    # Find kids of parents
    print("\n🔍 Matching kids with parents...")
    combined_df = find_kids_of_parents(parents_df, kids_df, backup_kids_df)
    
    # Get parent-kid mapping
    print("\n📊 Creating parent-kid mapping...")
    data_map = get_parent_kid_map(combined_df)
    
    # Calculate amounts paid
    print("\n💰 Calculating payments...")
    amount_map = calculate_months_paid(parents_df)
    
    # Get kids status
    print("\n📋 Getting kids status...")
    kids_status = get_all_kids_last_updates(KID_FILE, months)
    
    # Calculate kid payments
    print("\n🧮 Calculating kid payment statuses...")
    kid_payment_status = calculate_kid_payments(
        data_map, 
        amount_map,
        {row['kid_name']: {
            'allocated_amount': 0.0,
            'class': row['class'],
            'monthly_fee': get_monthly_fee_for_class(row['class']),
            'parent': row['parent_name']
        } for _, row in kids_df.iterrows()}
    )
    
    # Update Excel file
    print("\n" + "="*60)
    print("UPDATING EXCEL FILE")
    print("="*60 + "\n")
    
    output_file = update_excel_with_payments(
        kids_df=kids_df,
        kid_payment_status=kid_payment_status,
        kids_status=kids_status,
        months=months,
        kid_file=KID_FILE,
        output_file=OUTPUT_FILE
    )
    
    print(f"\n🎉 Process completed! Check '{output_file}' for results.")


if __name__ == "__main__":
    main()