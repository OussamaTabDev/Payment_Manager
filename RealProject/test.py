import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color
from openpyxl import load_workbook
import pandas as pd
from copy import copy

# [Your existing code here - all imports and functions]
# ... (keep all your existing code above)

def copy_cell_format(source_cell, target_cell):
    """Copy all formatting from source cell to target cell"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def extend_months_to_2years():
    """Extend the months list to cover 2 years instead of 1.5 years"""
    months_2years = [
        '9', '10', '11', '12', '1', '2', '3', '4', '5', '6', '7', '8',  # Year 1
        '9_next', '10_next', '11_next', '12_next', '1_next', '2_next', 
        '3_next', '4_next', '5_next', '6_next', '7_next', '8_next'  # Year 2
    ]
    return months_2years

def update_excel_with_payments(kids_df, kid_payment_status, kids_first_rows, kids_last_rows, 
                                output_file="kids_list_updated.xlsx"):
    """
    Update the Excel file with new payment statuses while preserving formatting
    """
    # Load the original workbook to preserve formatting
    wb = load_workbook("kids_list.xlsx")
    ws = wb.active
    
    # Extend months to 2 years
    months_extended = extend_months_to_2years()
    
    # Calculate how many new month columns to add
    original_month_count = len(months)
    new_month_count = len(months_extended)
    months_to_add = new_month_count - original_month_count
    
    # Get the position where month columns start (after kid_id, kid_name, parent_name)
    month_start_col = 4  # Column D (1-indexed)
    month_end_col = month_start_col + original_month_count - 1
    
    # Get a reference cell from existing months to copy style from
    reference_month_cell = ws.cell(row=4, column=month_start_col)
    
    # Insert new columns after the existing month columns
    if months_to_add > 0:
        for _ in range(months_to_add):
            ws.insert_cols(month_end_col + 1)
        
        # Find where 2026 starts (after the 24 months, which is 9_next to 8_next)
        year_2026_start_col = month_start_col + 24  # After 24 months
        
        # Merge cells for "2026" header (row 1)
        ws.merge_cells(start_row=1, start_column=year_2026_start_col, 
                      end_row=1, end_column=year_2026_start_col + months_to_add - 1)
        
        # Set "2026" text and copy style from "2025" cell
        year_2026_cell = ws.cell(row=1, column=year_2026_start_col)
        year_2026_cell.value = "2026"
        
        # Copy format from "2025" header
        year_2025_cell = ws.cell(row=1, column=month_start_col + 12)  # Where 2025 starts
        copy_cell_format(year_2025_cell, year_2026_cell)
        
        # Update headers for new month columns (row 2 and row 3)
        for i, month in enumerate(months_extended):
            col_idx = month_start_col + i
            
            # Row 2: Month numbers (9, 10, 11, 12, 1, 2, etc.)
            month_num_cell = ws.cell(row=2, column=col_idx)
            
            # Row 3: Month values (25, 25, 25, etc.)
            header_cell = ws.cell(row=3, column=col_idx)
            
            # For new columns (beyond original), copy format and set values
            if i >= original_month_count:
                # Get the month number (remove "_next" suffix for display)
                display_month = month.replace('_next', '')
                month_num_cell.value = display_month
                header_cell.value = month
                
                # Copy format from corresponding month in year 1
                # e.g., 9_next copies from 9, 10_next from 10, etc.
                source_col_offset = i % 12  # Cycle through 12 months
                source_col = month_start_col + source_col_offset
                
                source_month_num = ws.cell(row=2, column=source_col)
                source_header = ws.cell(row=3, column=source_col)
                
                copy_cell_format(source_month_num, month_num_cell)
                copy_cell_format(source_header, header_cell)
    
    # Get kids status info
    kids_status_dict = {}
    for _, row in kids_status.iterrows():
        kid_name = row['kid_name']
        kids_status_dict[kid_name] = {
            'last_month': row['last_month'],
            'last_text': row['last_text'],
            'last_color': row['last_color']
        }
    
    # Update parent names and add phone numbers
    # Find the parent_name column (should be column 3)
    parent_name_col = 3
    
    # Get the combined kids_parents data with phone numbers
    kids_parents_combined = find_kids_of_parrents(parents_df, kids_df)
    
    # Create a mapping of kid_id to parent info
    kid_to_parent_info = {}
    for _, row in kids_parents_combined.iterrows():
        kid_to_parent_info[row['kid_id']] = {
            'parent_name': row['parent_name'],
            'phone_number': row.get('phone_number', '')
        }
    
    # Insert a new column for phone numbers after parent_name (before months)
    # This will be column 4, and months will shift to column 5
    ws.insert_cols(parent_name_col + 1)
    
    # Update the header for phone number column
    phone_header_cell = ws.cell(row=3, column=parent_name_col + 1)
    phone_header_cell.value = "Phone Number"
    # Copy format from parent_name header
    parent_header_cell = ws.cell(row=3, column=parent_name_col)
    copy_cell_format(parent_header_cell, phone_header_cell)
    
    # Also update row 1 and row 2 for the phone number column
    phone_row1_cell = ws.cell(row=1, column=parent_name_col + 1)
    phone_row2_cell = ws.cell(row=2, column=parent_name_col + 1)
    copy_cell_format(ws.cell(row=1, column=parent_name_col), phone_row1_cell)
    copy_cell_format(ws.cell(row=2, column=parent_name_col), phone_row2_cell)
    
    # Adjust month_start_col since we inserted a column
    month_start_col = 5  # Now months start at column 5
    
    # Update reference cell after column insertion
    reference_month_cell = ws.cell(row=4, column=month_start_col)
    
    # Process each kid row (starting from row 4)
    start_row = 4
    for idx, kid_row in kids_df.iterrows():
        excel_row = start_row + idx
        kid_name = kid_row['kid_name']
        kid_id = kid_row['kid_id']
        
        if pd.isna(kid_name):
            continue
        
        # Update parent name and add phone number
        if kid_id in kid_to_parent_info:
            parent_info = kid_to_parent_info[kid_id]
            
            # Update parent name (column 3)
            parent_cell = ws.cell(row=excel_row, column=parent_name_col)
            parent_cell.value = parent_info['parent_name']
            
            # Add phone number (column 4)
            phone_cell = ws.cell(row=excel_row, column=parent_name_col + 1)
            phone_cell.value = parent_info['phone_number']
            # Copy format from parent cell
            copy_cell_format(parent_cell, phone_cell)
        
        # Get payment status for this kid
        payment_info = kid_payment_status.get(kid_name, {})
        
        if not payment_info:
            print(f"⚠️ No payment info found for {kid_name}, skipping updates")
            continue
        
        # Get kid's last status
        last_status = kids_status_dict.get(kid_name, {})
        last_month = last_status.get('last_month')
        last_color = last_status.get('last_color')
        last_text = last_status.get('last_text')
        
        # Find the index of the last updated month
        try:
            last_month_idx = months_extended.index(last_month) if last_month else -1
        except ValueError:
            last_month_idx = -1
        
        # Determine how many new months to fill based on payment
        months_paid = payment_info.get('months_paid', 0.0)
        full_months_paid = int(months_paid)
        monthly_fee = payment_info.get('monthly_fee', 0.0)
        
        # Determine the color for new entries
        new_color = payment_info.get('color', 'FF595959')
        new_status = payment_info.get('status', '')
        
        # Calculate extras
        extras = payment_info.get('extras', 0.0)
        extras_color = payment_info.get('extras_color', '#92d050')
        
        # Fill month columns
        for i, month in enumerate(months_extended):
            col_idx = month_start_col + i
            cell = ws.cell(row=excel_row, column=col_idx)
            
            # Keep old data as-is (before or at last_month_idx)
            if i <= last_month_idx:
                # Preserve existing formatting and value
                continue
            
            # Fill new months based on payment
            elif i <= last_month_idx + full_months_paid:
                # Put the monthly fee price
                if monthly_fee > 0:
                    cell.value = int(monthly_fee) if monthly_fee == int(monthly_fee) else monthly_fee
                else:
                    cell.value = ""
                
                # Copy style from reference cell
                copy_cell_format(reference_month_cell, cell)
                
                # Apply payment status color
                cell.fill = PatternFill(start_color=new_color.replace("#", ""), 
                                       end_color=new_color.replace("#", ""), 
                                       fill_type="solid")
            
            # Handle extras in the next month
            elif i == last_month_idx + full_months_paid + 1 and extras > 0:
                cell.value = extras if extras != int(extras) else int(extras)
                
                # Copy style from reference cell
                copy_cell_format(reference_month_cell, cell)
                
                # Apply extras color
                cell.fill = PatternFill(start_color=extras_color.replace("#", ""), 
                                       end_color=extras_color.replace("#", ""), 
                                       fill_type="solid")
            
            else:
                # Copy style and mark as unpaid (red)
                copy_cell_format(reference_month_cell, cell)
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFF0000", 
                                       end_color="FFFF0000", 
                                       fill_type="solid")
        
        print(f"✅ Updated {kid_name}: {full_months_paid} months paid (€{monthly_fee}/month), extras: €{extras}")
    
    # Save the updated workbook
    wb.save(output_file)
    print(f"\n✅ Excel file updated successfully: {output_file}")
    return output_file

# Add this at the end of your script, after calculate_kid_payments
print("\n" + "="*60)
print("UPDATING EXCEL FILE WITH NEW PAYMENT STATUS")
print("="*60 + "\n")

# Calculate kid payment statuses
kid_payment_status = calculate_kid_payments(data_map, amount_map, 
                                            {row['kid_name']: {
                                                'allocated_amount': 0.0,  # Start fresh or use previous
                                                'class': row['class'],
                                                'monthly_fee': get_monthly_fee_for_class(row['class']),
                                                'parent': row['parent_name']
                                            } for _, row in kids_df.iterrows()})

# Update the Excel file
output_file = update_excel_with_payments(
    kids_df=kids_df,
    kid_payment_status=kid_payment_status,
    kids_first_rows=kids_first_rows,
    kids_last_rows=kids_last_rows,
    output_file="kids_list_updated.xlsx"
)

print(f"\n🎉 Process completed! Check '{output_file}' for results.")