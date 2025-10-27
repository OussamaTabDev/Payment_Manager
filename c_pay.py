import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QPushButton, QLineEdit, 
                             QFileDialog, QTableWidget, QTableWidgetItem, 
                             QTabWidget, QSpinBox, QDoubleSpinBox, QTextEdit,
                             QMessageBox, QFrame, QScrollArea, QGroupBox,
                             QComboBox, QCheckBox)
from PyQt6.QtCore import Qt, QMimeData, pyqtSignal, QPropertyAnimation, QEasingCurve, QSettings
from PyQt6.QtGui import QDragEnterEvent, QDropEvent, QPalette, QColor, QFont, QIcon
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

class DragDropLabel(QLabel):
    """Custom label that accepts drag and drop for files"""
    fileDropped = pyqtSignal(str)
    
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setMinimumHeight(120)
        self.default_style = """
            QLabel {
                border: 2px dashed #666;
                border-radius: 10px;
                padding: 20px;
                background-color: rgba(100, 100, 100, 0.1);
                font-size: 14px;
            }
        """
        self.hover_style = """
            QLabel {
                border: 2px dashed #4CAF50;
                border-radius: 10px;
                padding: 20px;
                background-color: rgba(76, 175, 80, 0.1);
                font-size: 14px;
            }
        """
        self.setStyleSheet(self.default_style)
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(self.hover_style)
            
    def dragLeaveEvent(self, event):
        self.setStyleSheet(self.default_style)
        
    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            self.fileDropped.emit(files[0])
        self.setStyleSheet(self.default_style)

class PaymentTrackerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kids Payment Tracker 💰")
        self.setMinimumSize(1200, 800)
        
        # Data storage
        self.parents_df = None
        self.kids_df = None
        self.parents_file = None
        self.kids_file = None
        
        # Settings
        self.settings = QSettings("PaymentTracker", "KidsPayments")
        
        # Detect system theme
        self.dark_mode = self.is_system_dark_mode()
        
        # Month columns
        self.month_columns = ['January','February','March','April','May','June',
                             'July','August','September','October','November','December']
        
        self.init_ui()
        self.apply_theme()
        
    def is_system_dark_mode(self):
        """Detect if system is using dark mode"""
        palette = QApplication.palette()
        bg_color = palette.color(QPalette.ColorRole.Window)
        # If background is dark (luminance < 128), system is in dark mode
        return bg_color.lightness() < 128

    def browse_file(self, file_type):
        """Open a file dialog to select a file and load it."""
        file_name, _ = QFileDialog.getOpenFileName(
            self, 
            f"Select {file_type.title()} File", 
            "", 
            "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_name:
            self.load_file(file_name, file_type)
            
    def init_ui(self):
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = self.create_header()
        main_layout.addWidget(header)
        
        # Tab widget
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.TabPosition.North)
        
        # Tab 1: File Input & Settings
        input_tab = self.create_input_tab()
        self.tabs.addTab(input_tab, "📁 Files & Settings")
        
        # Tab 2: Process & Results
        results_tab = self.create_results_tab()
        self.tabs.addTab(results_tab, "📊 Process & Results")
        
        # Tab 3: Preview
        preview_tab = self.create_preview_tab()
        self.tabs.addTab(preview_tab, "👁️ Data Preview")
        
        main_layout.addWidget(self.tabs)
        
    def create_header(self):
        header_frame = QFrame()
        header_frame.setFrameShape(QFrame.Shape.StyledPanel)
        header_layout = QHBoxLayout(header_frame)
        
        # Title
        title = QLabel("🎓 Kids Payment Tracker")
        title.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Theme toggle
        self.theme_btn = QPushButton("🌙 Dark Mode")
        self.theme_btn.clicked.connect(self.toggle_theme)
        self.theme_btn.setMinimumHeight(40)
        header_layout.addWidget(self.theme_btn)
        
        # Help button
        help_btn = QPushButton("❓ Help")
        help_btn.clicked.connect(self.show_help)
        help_btn.setMinimumHeight(40)
        header_layout.addWidget(help_btn)
        
        return header_frame
        
    def create_input_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setSpacing(20)
        
        # Files section
        files_group = QGroupBox("📂 Input Files")
        files_layout = QHBoxLayout()  # Changed from QVBoxLayout to QHBoxLayout

        # --- Parents file section ---
        parents_layout = QVBoxLayout()
        parents_label = QLabel("<b>Parents Payments File</b> (Excel/CSV)")
        parents_layout.addWidget(parents_label)

        self.parents_drop = DragDropLabel("🖱️ Drag & Drop Excel/CSV file here\nor click below to browse")
        self.parents_drop.fileDropped.connect(lambda f: self.load_file(f, 'parents'))
        parents_layout.addWidget(self.parents_drop)

        parents_btn_layout = QHBoxLayout()
        parents_browse = QPushButton("📁 Browse Parents File")
        parents_browse.clicked.connect(lambda: self.browse_file('parents'))
        parents_btn_layout.addWidget(parents_browse)

        self.parents_status = QLabel("No file loaded")
        self.parents_status.setStyleSheet("color: #666; font-style: italic;")
        parents_btn_layout.addWidget(self.parents_status)
        parents_btn_layout.addStretch()

        parents_layout.addLayout(parents_btn_layout)
        parents_layout.addStretch()  # Optional: pushes content to top if group box is tall

        # --- Kids file section ---
        kids_layout = QVBoxLayout()
        kids_label = QLabel("<b>Kids List File</b> (Excel/CSV)")
        kids_layout.addWidget(kids_label)

        self.kids_drop = DragDropLabel("🖱️ Drag & Drop Excel/CSV file here\nor click below to browse")
        self.kids_drop.fileDropped.connect(lambda f: self.load_file(f, 'kids'))
        kids_layout.addWidget(self.kids_drop)

        kids_btn_layout = QHBoxLayout()
        kids_browse = QPushButton("📁 Browse Kids File")
        kids_browse.clicked.connect(lambda: self.browse_file('kids'))
        kids_btn_layout.addWidget(kids_browse)

        self.kids_status = QLabel("No file loaded")
        self.kids_status.setStyleSheet("color: #666; font-style: italic;")
        kids_btn_layout.addWidget(self.kids_status)
        kids_btn_layout.addStretch()

        kids_layout.addLayout(kids_btn_layout)
        kids_layout.addStretch()  # Optional: pushes content to top

        # Add both columns to the horizontal layout
        files_layout.addLayout(parents_layout)
        files_layout.addLayout(kids_layout)

        files_group.setLayout(files_layout)
        layout.addWidget(files_group)
        
        # Settings section
        settings_group = QGroupBox("⚙️ Settings")
        settings_layout = QVBoxLayout()
        
        # Monthly fee
        fee_layout = QHBoxLayout()
        fee_layout.addWidget(QLabel("💵 Monthly Fee per Kid:"))
        self.fee_input = QDoubleSpinBox()
        self.fee_input.setRange(0.01, 10000.0)
        self.fee_input.setValue(20.0)
        self.fee_input.setPrefix("$")
        self.fee_input.setDecimals(2)
        self.fee_input.setMinimumWidth(150)
        fee_layout.addWidget(self.fee_input)
        fee_layout.addStretch()
        
        hint_label = QLabel("💡 Tip: This is the monthly fee charged per child")
        hint_label.setStyleSheet("color: #888; font-size: 12px;")
        fee_layout.addWidget(hint_label)
        
        settings_layout.addLayout(fee_layout)
        
        # Output file name
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("📄 Output File Name:"))
        self.output_input = QLineEdit("updated_kids_list.xlsx")
        self.output_input.setMinimumWidth(300)
        output_layout.addWidget(self.output_input)
        output_layout.addStretch()
        settings_layout.addLayout(output_layout)
        
        # Apply styling option
        style_layout = QHBoxLayout()
        self.apply_style_check = QCheckBox("✨ Apply green highlighting to 'Paid' cells")
        self.apply_style_check.setChecked(True)
        style_layout.addWidget(self.apply_style_check)
        style_layout.addStretch()
        settings_layout.addLayout(style_layout)
        
        # Custom output location
        location_group = QGroupBox("📍 Output Location")
        location_layout = QVBoxLayout()
        
        # Default location option
        default_layout = QHBoxLayout()
        self.use_default_location = QCheckBox("Use default location (current folder)")
        self.use_default_location.setChecked(True)
        self.use_default_location.toggled.connect(self.toggle_custom_location)
        default_layout.addWidget(self.use_default_location)
        default_layout.addStretch()
        location_layout.addLayout(default_layout)
        
        # Custom location
        custom_layout = QHBoxLayout()
        custom_layout.addWidget(QLabel("Custom folder:"))
        self.custom_location_input = QLineEdit()
        self.custom_location_input.setPlaceholderText("Select a custom output folder...")
        self.custom_location_input.setEnabled(False)
        custom_layout.addWidget(self.custom_location_input)
        
        self.browse_location_btn = QPushButton("📁 Browse")
        self.browse_location_btn.clicked.connect(self.browse_output_location)
        self.browse_location_btn.setEnabled(False)
        custom_layout.addWidget(self.browse_location_btn)
        
        location_layout.addLayout(custom_layout)
        location_group.setLayout(location_layout)
        settings_layout.addWidget(location_group)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        # Process button (large and prominent)
        process_frame = QFrame()
        process_frame.setFrameShape(QFrame.Shape.StyledPanel)
        process_layout = QVBoxLayout(process_frame)
        
        self.main_process_btn = QPushButton("PROCESS PAYMENTS")
        self.main_process_btn.setMinimumHeight(70)
        self.main_process_btn.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        self.main_process_btn.clicked.connect(self.process_and_auto_save)
        self.main_process_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                           stop:0 #4CAF50, stop:1 #45a049);
                font-size: 16px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                           stop:0 #5CBF60, stop:1 #55b059);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                           stop:0 #3d8b40, stop:1 #357a38);
            }
        """)
        process_layout.addWidget(self.main_process_btn)
        
        hint = QLabel("💡 This will process payments and automatically save the result")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint.setStyleSheet("color: #888; font-size: 12px; font-style: italic;")
        process_layout.addWidget(hint)
        
        layout.addWidget(process_frame)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        layout.addStretch()
        scroll.setWidget(container)
        return scroll
        
    def create_results_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        
        # Process button
        process_layout = QHBoxLayout()
        self.process_btn = QPushButton("🚀 Process Payments")
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        self.process_btn.clicked.connect(self.process_payments)
        process_layout.addWidget(self.process_btn)
        layout.addLayout(process_layout)
        
        # Results text area
        results_label = QLabel("<b>📋 Processing Results:</b>")
        layout.addWidget(results_label)
        
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        self.results_text.setFont(QFont("Courier", 10))
        layout.addWidget(self.results_text)
        
        # Action buttons
        action_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("💾 Save Results")
        self.save_btn.clicked.connect(self.save_results)
        self.save_btn.setEnabled(False)
        self.save_btn.setMinimumHeight(40)
        action_layout.addWidget(self.save_btn)
        
        self.open_btn = QPushButton("📂 Open Output Folder")
        self.open_btn.clicked.connect(self.open_output_folder)
        self.open_btn.setEnabled(False)
        self.open_btn.setMinimumHeight(40)
        action_layout.addWidget(self.open_btn)
        
        action_layout.addStretch()
        layout.addLayout(action_layout)
        
        return widget
    
    def toggle_custom_location(self, checked):
        """Enable/disable custom location inputs"""
        self.custom_location_input.setEnabled(not checked)
        self.browse_location_btn.setEnabled(not checked)
        
    def browse_output_location(self):
        """Browse for custom output folder"""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.custom_location_input.setText(folder)
        
    def create_preview_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Preview selector
        selector_layout = QHBoxLayout()
        selector_layout.addWidget(QLabel("Preview:"))
        self.preview_combo = QComboBox()
        self.preview_combo.addItems(["Parents Payments", "Kids List", "Updated Kids List"])
        self.preview_combo.currentTextChanged.connect(self.update_preview)
        selector_layout.addWidget(self.preview_combo)
        selector_layout.addStretch()
        
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self.update_preview)
        selector_layout.addWidget(refresh_btn)
        
        layout.addLayout(selector_layout)
        
        # Table
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        layout.addWidget(self.preview_table)
        
        return widget
        
    def toggle_custom_location(self, checked):
        """Enable/disable custom location inputs"""
        self.custom_location_input.setEnabled(not checked)
        self.browse_location_btn.setEnabled(not checked)
        
    def browse_output_location(self):
        """Browse for custom output folder"""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.custom_location_input.setText(folder)
            
    def process_and_auto_save(self):
        """Process payments and automatically save, then switch to results tab"""
        if self.parents_df is None or self.kids_df is None:
            QMessageBox.warning(self, "Missing Data", "Please load both parent and kids files first!")
            return
            
        try:
            # Switch to results tab
            self.tabs.setCurrentIndex(1)
            
            # Process payments
            self.results_text.clear()
            self.results_text.append("🔄 Starting payment processing...\n")
            
            monthly_fee = self.fee_input.value()
            
            # Find parent-kid relationships
            self.results_text.append("👨‍👩‍👧‍👦 Finding parent-kid relationships...")
            parent_kid_map = self.find_kids_of_parents(self.parents_df, self.kids_df)
            
            # Listing kids from less paid months to more 
            listed_parent_kid_map = {}
            listed_parent_kid_map = self.listing_parent_kid_map(parent_kid_map , self.kids_df )

            for parent, kids in listed_parent_kid_map.items():
                self.results_text.append(f"  • {parent} → {', '.join(kids)}")
            self.results_text.append("")
            
            # Calculate months paid
            self.results_text.append("💰 Calculating payments...")
            kids_months_paid = self.calculate_months_paid(self.parents_df, listed_parent_kid_map, monthly_fee)
            
            for kid, months in kids_months_paid.items():
                self.results_text.append(f"  • {kid}: {months} months paid")
            self.results_text.append("")
            
            # Update dataframe
            self.results_text.append("📝 Updating kids payment records...")
            self.updated_kids_df = self.update_kids_months_paid(kids_months_paid, self.kids_df.copy())
            
            self.results_text.append("\n✅ Processing complete!")
            
            # Auto-save
            self.results_text.append("\n💾 Auto-saving results...")
            self.auto_save_results()
            
            self.save_btn.setEnabled(True)
            self.update_preview()
            
        except Exception as e:
            QMessageBox.critical(self, "Processing Error", f"Error during processing:\n{str(e)}")
            self.results_text.append(f"\n❌ Error: {str(e)}")


    def auto_save_results(self):
        """Automatically save results to specified location"""
        try:
            # Determine output path
            output_filename = self.output_input.text()
            
            if self.use_default_location.isChecked():
                output_file = output_filename
            else:
                custom_folder = self.custom_location_input.text()
                if not custom_folder:
                    raise Exception("Custom location not specified")
                output_file = os.path.join(custom_folder, output_filename)
            
            # Save to Excel
            self.updated_kids_df.to_excel(output_file, index=False)
            
            # Apply styling if checked
            if self.apply_style_check.isChecked():
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active
                
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                headers = [cell.value for cell in ws[1]]
                month_col_indices = [idx for idx, h in enumerate(headers, start=1) if h in self.month_columns]
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for col_idx in month_col_indices:
                        cell = row[col_idx - 1]
                        if cell.value == "Paid":
                            cell.fill = green_fill
                            
                wb.save(output_file)
            
            self.results_text.append(f"✅ File saved successfully to:\n   {os.path.abspath(output_file)}")
            self.open_btn.setEnabled(True)
            
            # Show success message
            QMessageBox.information(
                self, 
                "Success! 🎉", 
                f"Processing complete!\n\nFile saved to:\n{os.path.abspath(output_file)}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save file:\n{str(e)}")
            self.results_text.append(f"\n❌ Save Error: {str(e)}")
        file_name, _ = QFileDialog.getOpenFileName(
            self, f"Select {file_type.title()} File", 
            "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_name:
            self.load_file(file_name, file_type)
            
    def load_file(self, file_path, file_type):
        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
                
            if file_type == 'parents':
                self.parents_df = df
                self.parents_file = file_path
                self.parents_status.setText(f"✅ Loaded: {os.path.basename(file_path)} ({len(df)} rows)")
                self.parents_status.setStyleSheet("color: green;")
                self.parents_drop.setText(f"✅ {os.path.basename(file_path)}\n{len(df)} rows loaded")
            else:
                self.kids_df = df
                self.kids_file = file_path
                self.kids_status.setText(f"✅ Loaded: {os.path.basename(file_path)} ({len(df)} rows)")
                self.kids_status.setStyleSheet("color: green;")
                self.kids_drop.setText(f"✅ {os.path.basename(file_path)}\n{len(df)} rows loaded")
                
            self.update_preview()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load file:\n{str(e)}")
    
    def process_and_auto_save(self):
        """Process payments and automatically save, then switch to results tab"""
        if self.parents_df is None or self.kids_df is None:
            QMessageBox.warning(self, "Missing Data", "Please load both parent and kids files first!")
            return
            
        try:
            # Switch to results tab
            self.tabs.setCurrentIndex(1)
            
            # Process payments
            self.results_text.clear()
            self.results_text.append("🔄 Starting payment processing...\n")
            
            monthly_fee = self.fee_input.value()
            
            # Find parent-kid relationships
            self.results_text.append("👨‍👩‍👧‍👦 Finding parent-kid relationships...")
            parent_kid_map = self.find_kids_of_parents(self.parents_df, self.kids_df)
            
            # Listing kids from less paid months to more 
            listed_parent_kid_map = {}
            listed_parent_kid_map = self.listing_parent_kid_map(parent_kid_map , self.kids_df )

            for parent, kids in listed_parent_kid_map.items():
                self.results_text.append(f"  • {parent} → {', '.join(kids)}")
            self.results_text.append("")
            
            # Calculate months paid
            self.results_text.append("💰 Calculating payments...")
            kids_months_paid = self.calculate_months_paid(self.parents_df, listed_parent_kid_map, monthly_fee)
            
            for kid, months in kids_months_paid.items():
                self.results_text.append(f"  • {kid}: {months} months paid")
            self.results_text.append("")
            
            # Update dataframe
            self.results_text.append("📝 Updating kids payment records...")
            self.updated_kids_df = self.update_kids_months_paid(kids_months_paid, self.kids_df.copy())
            
            self.results_text.append("\n✅ Processing complete!")
            
            # Auto-save
            self.results_text.append("\n💾 Auto-saving results...")
            self.auto_save_results()
            
            self.save_btn.setEnabled(True)
            self.update_preview()
            
        except Exception as e:
            QMessageBox.critical(self, "Processing Error", f"Error during processing:\n{str(e)}")
            self.results_text.append(f"\n❌ Error: {str(e)}")
            
    def auto_save_results(self):
        """Automatically save results to specified location"""
        try:
            # Determine output path
            output_filename = self.output_input.text()
            
            if self.use_default_location.isChecked():
                output_file = output_filename
            else:
                custom_folder = self.custom_location_input.text()
                if not custom_folder:
                    raise Exception("Custom location not specified")
                output_file = os.path.join(custom_folder, output_filename)
            
            # Save to Excel
            self.updated_kids_df.to_excel(output_file, index=False)
            
            # Apply styling if checked
            if self.apply_style_check.isChecked():
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active
                
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                headers = [cell.value for cell in ws[1]]
                month_col_indices = [idx for idx, h in enumerate(headers, start=1) if h in self.month_columns]
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for col_idx in month_col_indices:
                        cell = row[col_idx - 1]
                        if cell.value == "Paid":
                            cell.fill = green_fill
                            
                wb.save(output_file)
            
            self.results_text.append(f"✅ File saved successfully to:\n   {os.path.abspath(output_file)}")
            self.open_btn.setEnabled(True)
            
            # Show success message
            QMessageBox.information(
                self, 
                "Success! 🎉", 
                f"Processing complete!\n\nFile saved to:\n{os.path.abspath(output_file)}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save file:\n{str(e)}")
            self.results_text.append(f"\n❌ Save Error: {str(e)}")
            
    def process_payments(self):
        if self.parents_df is None or self.kids_df is None:
            QMessageBox.warning(self, "Missing Data", "Please load both parent and kids files first!")
            return
            
        try:
            self.results_text.clear()
            self.results_text.append("🔄 Starting payment processing...\n")
            
            monthly_fee = self.fee_input.value()
            
            # Find parent-kid relationships
            self.results_text.append("👨‍👩‍👧‍👦 Finding parent-kid relationships...")
            parent_kid_map = self.find_kids_of_parents(self.parents_df, self.kids_df)
            
            for parent, kids in parent_kid_map.items():
                self.results_text.append(f"  • {parent} → {', '.join(kids)}")
            self.results_text.append("")
            
            # Calculate months paid
            self.results_text.append("💰 Calculating payments...")
            kids_months_paid = self.calculate_months_paid(self.parents_df, parent_kid_map, monthly_fee)
            
            for kid, months in kids_months_paid.items():
                self.results_text.append(f"  • {kid}: {months} months paid")
            self.results_text.append("")
            
            # Update dataframe
            self.results_text.append("📝 Updating kids payment records...")
            self.updated_kids_df = self.update_kids_months_paid(kids_months_paid, self.kids_df.copy())
            
            self.results_text.append("\n✅ Processing complete!")
            self.save_btn.setEnabled(True)
            
            self.update_preview()
            
        except Exception as e:
            QMessageBox.critical(self, "Processing Error", f"Error during processing:\n{str(e)}")
            self.results_text.append(f"\n❌ Error: {str(e)}")
            
    def find_kids_of_parents(self, parents_df, kids_df):
        distinct_parents = parents_df['parents_name'].dropna().unique()
        distinct_kids = kids_df['kid_name'].dropna().unique()
        parent_kid_map = {}
        
        for parent in distinct_parents:
            last_name_parent = parent.split()[-1]
            matched_kids = [kid for kid in distinct_kids if last_name_parent.lower() == kid.split()[-1].lower()]
            # matched_kids = [kid for kid in distinct_kids if last_name_parent == kid.split()[-1]]
            if matched_kids:
                parent_kid_map[parent] = matched_kids
                
        return parent_kid_map

    def listing_parent_kid_map(self, parent_kid_map, kids_df):
        listed_parent_kid_map = {}
        
        for parent, kids in parent_kid_map.items():
            kids_months = {}
            for kid in kids:
                kid_row = kids_df[kids_df['kid_name'] == kid]
                if not kid_row.empty:
                    paid_months = sum(1 for month in self.month_columns if str(kid_row.iloc[0][month]).strip().lower() == 'paid')
                    kids_months[kid] = paid_months
                else:
                    kids_months[kid] = 0
                    
            # Sort kids by months paid (ascending)
            sorted_kids = sorted(kids_months.items(), key=lambda x: x[1])
            listed_parent_kid_map[parent] = [kid for kid, _ in sorted_kids]
            
        return listed_parent_kid_map

    def getting_amount_from_string(self, amount_str):
        try:
            amount = int(''.join(filter(str.isdigit, str(amount_str))))
            return amount
        except:
            return 0.0
            
    def calculate_months_paid(self, parents_df, parent_kid_map, monthly_fee):
        parents_amount = dict(zip(
            parents_df['parents_name'], 
            (parents_df['amount'].apply(self.getting_amount_from_string) / monthly_fee).round().astype(int)
        ))
        
        kids_months_paid = {}
        
        for parent, kids in parent_kid_map.items():
            if parent in parents_amount:
                months_paid = parents_amount[parent]
                num_kids = len(kids)
                months_per_kid = months_paid // num_kids if num_kids > 0 else 0
                months_module = months_paid % num_kids if num_kids > 0 else 0
                
                for kid in kids:
                    kids_months_paid[kid] = months_per_kid + (1 if months_module > 0 else 0)
                    months_module -= 1 if months_module > 0 else 0
                    
        return kids_months_paid
        
    def update_kids_months_paid(self, kids_months_paid, kids_df):
        updated_df = kids_df.copy()
        
        for kid_name, months_to_pay in kids_months_paid.items():
            kid_mask = updated_df['kid_name'] == kid_name
            
            if kid_mask.any():
                row_idx = updated_df[kid_mask].index[0]
                updated_row = self.mark_paid(updated_df.loc[row_idx], months_to_pay)
                updated_df.loc[row_idx] = updated_row
                
        return updated_df
        
    def mark_paid(self, row, months_to_pay):
        start_idx = 0
        for i, month_col in enumerate(self.month_columns):
            if pd.isna(row[month_col]) or row[month_col] == '':
                start_idx = i
                break
        else:
            return row
            
        for i in range(start_idx, min(start_idx + months_to_pay, len(self.month_columns))):
            row[self.month_columns[i]] = "Paid"
            
        return row
        
    def save_results(self):
        if not hasattr(self, 'updated_kids_df'):
            QMessageBox.warning(self, "No Results", "Please process payments first!")
            return
            
        try:
            output_file = self.output_input.text()
            self.updated_kids_df.to_excel(output_file, index=False)
            
            if self.apply_style_check.isChecked():
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active
                
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                headers = [cell.value for cell in ws[1]]
                month_col_indices = [idx for idx, h in enumerate(headers, start=1) if h in self.month_columns]
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for col_idx in month_col_indices:
                        cell = row[col_idx - 1]
                        if cell.value == "Paid":
                            cell.fill = green_fill
                            
                wb.save(output_file)
                
            QMessageBox.information(self, "Success", f"✅ File saved successfully:\n{output_file}")
            self.open_btn.setEnabled(True)
            self.results_text.append(f"\n💾 Saved to: {output_file}")
            
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save file:\n{str(e)}")
            
    def open_output_folder(self):
        """Open the folder containing the output file"""
        output_filename = self.output_input.text()
        
        if self.use_default_location.isChecked():
            folder = os.path.dirname(os.path.abspath(output_filename)) if os.path.dirname(output_filename) else os.getcwd()
        else:
            folder = self.custom_location_input.text() or os.getcwd()
            
        if sys.platform == "win32":
            os.startfile(folder)
        elif sys.platform == "darwin":
            os.system(f'open "{folder}"')
        else:
            os.system(f'xdg-open "{folder}"')
        
    def update_preview(self):
        selection = self.preview_combo.currentText()
        
        df = None
        if selection == "Parents Payments" and self.parents_df is not None:
            df = self.parents_df
        elif selection == "Kids List" and self.kids_df is not None:
            df = self.kids_df
        elif selection == "Updated Kids List" and hasattr(self, 'updated_kids_df'):
            df = self.updated_kids_df
            
        if df is not None:
            self.preview_table.setRowCount(len(df))
            self.preview_table.setColumnCount(len(df.columns))
            self.preview_table.setHorizontalHeaderLabels(df.columns.tolist())
            
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    item = QTableWidgetItem(str(df.iloc[i, j]))
                    if hasattr(self, 'updated_kids_df') and selection == "Updated Kids List":
                        if str(df.iloc[i, j]) == "Paid":
                            item.setBackground(QColor("#005C12"))
                            # item.setBackground(QColor("#C6EFCE"))
                    self.preview_table.setItem(i, j, item)
                    
            self.preview_table.resizeColumnsToContents()
        else:
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            
    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.apply_theme()
        
    def apply_theme(self):
        if self.dark_mode:
            self.setStyleSheet("""
                QMainWindow, QWidget {
                    background-color: #1e1e1e;
                    color: #e0e0e0;
                }
                QGroupBox {
                    border: 2px solid #444;
                    border-radius: 8px;
                    margin-top: 12px;
                    padding-top: 10px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    padding: 0 5px;
                }
                QPushButton {
                    background-color: #0d47a1;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #1565c0;
                }
                QPushButton:pressed {
                    background-color: #003c8f;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
                    background-color: #2d2d2d;
                    border: 1px solid #444;
                    border-radius: 4px;
                    padding: 5px;
                    color: #e0e0e0;
                }
                QTextEdit {
                    background-color: #2d2d2d;
                    border: 1px solid #444;
                    border-radius: 4px;
                    color: #e0e0e0;
                }
                QTableWidget {
                    background-color: #2d2d2d;
                    border: 1px solid #444;
                    gridline-color: #444;
                    color: #e0e0e0;
                }
                QHeaderView::section {
                    background-color: #383838;
                    color: #e0e0e0;
                    padding: 5px;
                    border: 1px solid #444;
                }
                QTabWidget::pane {
                    border: 1px solid #444;
                }
                QTabBar::tab {
                    background-color: #2d2d2d;
                    color: #e0e0e0;
                    padding: 10px 20px;
                    border: 1px solid #444;
                }
                QTabBar::tab:selected {
                    background-color: #0d47a1;
                }
            """)
            self.theme_btn.setText("☀️ Light Mode")
        else:
            self.setStyleSheet("""
                QMainWindow, QWidget {
                    background-color: #f5f5f5;
                    color: #333;
                }
                QGroupBox {
                    border: 2px solid #ddd;
                    border-radius: 8px;
                    margin-top: 12px;
                    padding-top: 10px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    padding: 0 5px;
                }
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #42A5F5;
                }
                QPushButton:pressed {
                    background-color: #1976D2;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
                    background-color: white;
                    border: 1px solid #ccc;
                    border-radius: 4px;
                    padding: 5px;
                }
                QTextEdit {
                    background-color: white;
                    border: 1px solid #ccc;
                    border-radius: 4px;
                }
                QTableWidget {
                    background-color: white;
                    border: 1px solid #ccc;
                    gridline-color: #ddd;
                }
                QHeaderView::section {
                    background-color: #e0e0e0;
                    padding: 5px;
                    border: 1px solid #ccc;
                }
                QTabWidget::pane {
                    border: 1px solid #ccc;
                }
                QTabBar::tab {
                    background-color: #e0e0e0;
                    padding: 10px 20px;
                    border: 1px solid #ccc;
                }
                QTabBar::tab:selected {
                    background-color: #2196F3;
                    color: white;
                }
            """)
            self.theme_btn.setText("🌙 Dark Mode")
            
    def show_help(self):
        help_text = """
        <h2>🎓 Kids Payment Tracker - Help Guide</h2>
        
        <h3>📁 How to Use:</h3>
        <ol>
            <li><b>Load Files:</b> Drag & drop or browse for your Excel/CSV files
                <ul>
                    <li>Parents file should have 'parents_name' and 'amount' columns</li>
                    <li>Kids file should have 'kid_name' column and month columns</li>
                </ul>
            </li>
            <li><b>Configure Settings:</b> Set monthly fee and output file name</li>
            <li><b>Process:</b> Click "Process Payments" to calculate</li>
            <li><b>Save:</b> Review results and save the updated file</li>
        </ol>
        
        <h3>✨ Features:</h3>
        <ul>
            <li>🖱️ Drag & drop file support</li>
            <li>🌙 Light/Dark mode toggle</li>
            <li>👁️ Live data preview</li>
            <li>✅ Automatic payment distribution across kids</li>
            <li>🎨 Green highlighting for paid months</li>
        </ul>
        
        <h3>💡 Tips:</h3>
        <ul>
            <li>The app matches kids to parents by last name</li>
            <li>Payments are distributed evenly among siblings</li>
            <li>Any remainder months go to the first kids in the list</li>
        </ul>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("Help")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(help_text)
        msg.exec()

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = PaymentTrackerApp()
    window.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()