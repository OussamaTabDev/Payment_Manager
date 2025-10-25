import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd

# Load the Excel file (or CSV)
parents_df = pd.read_excel("parents_payments.xlsx")
kids_df = pd.read_excel("kids_list.xlsx")

monthly_fee_per_kid = 20.0  # Example monthly fee per kid

# List of month columns - renamed to avoid conflict
month_columns = ['January','February','March','April','May','June','July','August','September','October','November','December']

print("Data loaded successfully.")

print("Showing data samples:")
print("Parents Payments Data:")
print(parents_df.head())
print("\nKids Full Names Data:")
print(kids_df.head())


# find parents names in kids full names
def find_kids_of_parrents(parents_df, kids_df):
    distinct_parents = parents_df['parents_name'].dropna().unique()
    distinct_kids = kids_df['kid_name'].dropna().unique()
    # put each parents name as key and value as list of kids names that contain the parents name
    parent_kid_map = {}
    for parent in distinct_parents:
        last_name_parent = parent.split()[-1]
        # print(f"Parent: {parent}, Last Name: {last_name_parent}")
        matched_kids = [kid for kid in distinct_kids if last_name_parent == kid.split()[-1]]
        if matched_kids:
            parent_kid_map[parent] = matched_kids
            # print(f"Parent: {parent} has kids: {matched_kids}")
    return parent_kid_map

def getting_mount_from_string(amount_str):
    try:
        amount = int(''.join(filter(str.isdigit, amount_str)))
        return amount
    except:
        return 0.0

def calculate_months_paid(parents_df = parents_df, parent_kid_map = {} , monthly_fee_per_kid=20.0):
    parents_mount = dict(zip(parents_df['parents_name'], (parents_df['amount'].apply(getting_mount_from_string)/monthly_fee_per_kid).round().astype(int)))
    print("Parents mount calculated:")
    print(parents_mount)

    # Calculate months paid for each kid using the parent_kid_map
    kids_months_paid = {}
    
    for parent, kids in parent_kid_map.items():
        if parent in parents_mount:
            months_paid = parents_mount[parent]
            num_kids = len(kids)
            months_per_kid = months_paid // num_kids if num_kids > 0 else 0
            months_module = months_paid % num_kids if num_kids > 0 else 0

            for kid in kids:
                kids_months_paid[kid] = months_per_kid + (1 if months_module > 0 else 0) # Distribute remainder months
                months_module -= 1 if months_module > 0 else 0
            # Distribute remainder months to the first few kids
    return kids_months_paid



def update_kids_months_paid_pd(kids_months_paid, kids_df):
    print("Updating kids months paid in DataFrame...")
    
    # Create a copy to avoid modifying the original during iteration
    updated_df = kids_df.copy()
    
    for kid_name, months_to_pay in kids_months_paid.items():
        # Find the row index for this kid
        kid_mask = updated_df['kid_name'] == kid_name
        
        if kid_mask.any():
            # Get the row index
            row_idx = updated_df[kid_mask].index[0]
            
            # Apply mark_paid to this specific row
            updated_row = mark_paid(updated_df.loc[row_idx], months_to_pay)
            updated_df.loc[row_idx] = updated_row
            print(f"Updated {kid_name}: {months_to_pay} months paid")
        else:
            print(f"Warning: Kid '{kid_name}' not found in dataframe")

    print(updated_df.head())
    return updated_df

def mark_paid(row, months_to_pay):
    """Mark the next unpaid months as paid"""
    # Find the fidef mark_paid(row, months_to_pay):
    """Mark the next unpaid months as paid"""
    # Find the first unpaid month
    start_idx = 0
    for i, month_col in enumerate(month_columns):  # Changed variable name from 'month' to 'month_col'
        if pd.isna(row[month_col]) or row[month_col] == '':
            start_idx = i
            break
    else:
        # All months are already paid
        return row
    
    # Mark the next 'months_to_pay' months as paid
    for i in range(start_idx, min(start_idx + months_to_pay, len(month_columns))):
        row[month_columns[i]] = "Paid"
    
    return row

print("Finding distinct parents...")
parent_kid_map = find_kids_of_parrents(parents_df, kids_df)

print("Print parents to their kids:")
for parent, kids in parent_kid_map.items():
    print(f"Parent: {parent} -> Kids: {', '.join(kids)}")

kids_months_paid = calculate_months_paid(parents_df , parent_kid_map , monthly_fee_per_kid )

print("Kids months paid calculated:")
for kid, months in kids_months_paid.items():
    print(f"Kid: {kid} -> Months Paid: {months}")

# Update the dataframe and assign it back
kids_df = update_kids_months_paid_pd(kids_months_paid, kids_df)

print("\nUpdated kids dataframe:")
print(kids_df.head())

# Save the updated dataframe back to Excel if needed
# kids_df.to_excel("updated_kids_list.xlsx", index=False)


# Save the updated dataframe first
output_file = "updated_kids_list.xlsx"
kids_df.to_excel(output_file, index=False)

# Now apply styling with openpyxl
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# Define green fill (you can also use Font color if preferred)
from openpyxl.styles import PatternFill
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green (like Excel's "Good" style)

# Get the header row to find column indices for month columns
headers = [cell.value for cell in ws[1]]  # First row

# Find column indices (1-based) for each month column
month_col_indices = []
for col_idx, header in enumerate(headers, start=1):
    if header in month_columns:
        month_col_indices.append(col_idx)

# Iterate through rows (skip header)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for col_idx in month_col_indices:
        cell = row[col_idx - 1]  # row is 0-based tuple
        if cell.value == "Paid":
            cell.fill = green_fill
            # Optional: make text bold or change font color
            # cell.font = Font(color="006100")  # Dark green text

# Save the styled workbook
wb.save(output_file)
print(f"\n✅ Excel file saved with green 'Paid' cells: {output_file}")