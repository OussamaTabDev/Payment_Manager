import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QTextEdit, QMessageBox,
    QDoubleSpinBox, QGroupBox, QSizePolicy , QCheckBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont


class Worker(QThread):
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str)

    def __init__(self, parents_file, kids_file, output_file, monthly_fee, month_columns):
        super().__init__()
        self.parents_file = parents_file
        self.kids_file = kids_file
        self.output_file = output_file
        self.monthly_fee = monthly_fee
        self.month_columns = month_columns

    def run(self):
        try:
            self.log_signal.emit("🔄 Loading data...")
            parents_df = pd.read_excel(self.parents_file)
            kids_df = pd.read_excel(self.kids_file)

        
            def find_kids_of_parents(parents_df, kids_df):
                distinct_parents = parents_df['parents_name'].dropna().unique()
                distinct_kids = kids_df['kid_name'].dropna().unique()
                parent_kid_map = {}
                for parent in distinct_parents:
                    last_name_parent = parent.split()[-1]
                    matched_kids = [kid for kid in distinct_kids if last_name_parent == kid.split()[-1]]
                    if matched_kids:
                        parent_kid_map[parent] = matched_kids
                return parent_kid_map

            def get_amount_from_string(amount_str):
                try:
                    return int(''.join(filter(str.isdigit, str(amount_str))))
                except:
                    return 0

            def calculate_months_paid(parents_df, parent_kid_map, monthly_fee):
                parents_amount = dict(zip(
                    parents_df['parents_name'],
                    parents_df['amount'].apply(get_amount_from_string)
                ))
                kids_months_paid = {}
                for parent, kids in parent_kid_map.items():
                    total_paid = parents_amount.get(parent, 0)
                    months_paid = int(round(total_paid / monthly_fee)) if monthly_fee > 0 else 0
                    num_kids = len(kids)
                    if num_kids == 0:
                        continue
                    base_months = months_paid // num_kids
                    remainder = months_paid % num_kids
                    for i, kid in enumerate(kids):
                        kids_months_paid[kid] = base_months + (1 if i < remainder else 0)
                return kids_months_paid

            def mark_paid(row, months_to_pay, month_columns):
                start_idx = 0
                for i, col in enumerate(month_columns):
                    if pd.isna(row[col]) or row[col] == '':
                        start_idx = i
                        break
                else:
                    return row
                for i in range(start_idx, min(start_idx + months_to_pay, len(month_columns))):
                    row[month_columns[i]] = "Paid"
                return row

            parent_kid_map = find_kids_of_parents(parents_df, kids_df)
            kids_months_paid = calculate_months_paid(parents_df, parent_kid_map, self.monthly_fee)

            updated_df = kids_df.copy()
            for kid, months in kids_months_paid.items():
                mask = updated_df['kid_name'] == kid
                if mask.any():
                    idx = updated_df[mask].index[0]
                    updated_df.loc[idx] = mark_paid(updated_df.loc[idx], months, self.month_columns)

            updated_df.to_excel(self.output_file, index=False)

            # Apply green styling
            wb = openpyxl.load_workbook(self.output_file)
            ws = wb.active
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            headers = [cell.value for cell in ws[1]]
            month_col_indices = [i+1 for i, h in enumerate(headers) if h in self.month_columns]

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in month_col_indices:
                    cell = row[col_idx - 1]
                    if cell.value == "Paid":
                        cell.fill = green_fill

            wb.save(self.output_file)
            self.log_signal.emit("✅ Done! File saved with green 'Paid' cells.")
            self.finished_signal.emit(self.output_file)

        except Exception as e:
            self.log_signal.emit(f"❌ Error: {str(e)}")
            self.finished_signal.emit("")


class PaymentTrackerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kids Payment Tracker")
        self.resize(800, 600)
        self.month_columns = [
            'January','February','March','April','May','June',
            'July','August','September','October','November','December'
        ]
        self.init_ui()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        # Title
        title = QLabel("Kids Payment Tracker")
        title.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # File Selection Group
        file_group = QGroupBox("📁 Input & Output Files")
        file_layout = QVBoxLayout()
        
        # Parents file
        self.parents_line = QLineEdit()
        self.parents_line.setPlaceholderText("Select Parents Payments Excel file...")
        parents_btn = QPushButton("Browse Parents File")
        parents_btn.clicked.connect(lambda: self.browse_file(self.parents_line, "Excel Files (*.xlsx)"))
        file_layout.addWidget(QLabel("Parents Payments File:"))
        file_layout.addWidget(self.parents_line)
        file_layout.addWidget(parents_btn)

        # Kids file
        self.kids_line = QLineEdit()
        self.kids_line.setPlaceholderText("Select Kids List Excel file...")
        kids_btn = QPushButton("Browse Kids File")
        kids_btn.clicked.connect(lambda: self.browse_file(self.kids_line, "Excel Files (*.xlsx)"))
        file_layout.addWidget(QLabel("Kids List File:"))
        file_layout.addWidget(self.kids_line)
        file_layout.addWidget(kids_btn)

        # Output option
        self.custom_output_checkbox = QCheckBox("Choose custom output location (default: next to Kids file)")
        self.custom_output_checkbox.stateChanged.connect(self.toggle_output_controls)
        file_layout.addWidget(self.custom_output_checkbox)

        self.output_line = QLineEdit()
        self.output_line.setPlaceholderText("Output will be saved next to Kids file (e.g., kids_list_updated.xlsx)")
        self.output_line.setEnabled(False)
        self.output_btn = QPushButton("Set Output File")
        self.output_btn.clicked.connect(self.set_output_file)
        self.output_btn.setEnabled(False)

        file_layout.addWidget(self.output_line)
        file_layout.addWidget(self.output_btn)

        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # Settings
        settings_group = QGroupBox("⚙️ Settings")
        settings_layout = QHBoxLayout()
        settings_layout.addWidget(QLabel("Monthly Fee per Kid:"))
        self.fee_spin = QDoubleSpinBox()
        self.fee_spin.setRange(0.1, 10000)
        self.fee_spin.setValue(20.0)
        self.fee_spin.setDecimals(2)
        settings_layout.addWidget(self.fee_spin)
        settings_layout.addStretch()
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)

        # Process Button
        self.process_btn = QPushButton("🚀 Process Payments")
        self.process_btn.setStyleSheet("QPushButton { font-size: 14px; padding: 10px; }")
        self.process_btn.clicked.connect(self.start_processing)
        layout.addWidget(self.process_btn)

        # Log Console
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Courier", 10))
        layout.addWidget(QLabel("📋 Log:"))
        layout.addWidget(self.log_text)

        self.log_text.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

    def toggle_output_controls(self, state):
        enabled = state == Qt.CheckState.Checked.value
        self.output_line.setEnabled(enabled)
        self.output_btn.setEnabled(enabled)
        if not enabled:
            self.output_line.clear()
            self.output_line.setPlaceholderText("Output will be saved next to Kids file (e.g., kids_list_updated.xlsx)")

    def set_output_file(self):
        file, _ = QFileDialog.getSaveFileName(
            self, "Save Output File", "", "Excel Files (*.xlsx)"
        )
        if file:
            if not file.endswith(".xlsx"):
                file += ".xlsx"
            self.output_line.setText(file)

    

    def browse_file(self, line_edit, file_filter):
        file, _ = QFileDialog.getOpenFileName(self, "Select File", "", file_filter)
        if file:
            line_edit.setText(file)

    

    def start_processing(self):
        parents = self.parents_line.text()
        kids = self.kids_line.text()
        fee = self.fee_spin.value()

        if not all([parents, kids]):
            QMessageBox.warning(self, "Missing Input", "Please select both Parents and Kids files.")
            return
        if not os.path.exists(parents):
            QMessageBox.warning(self, "File Missing", "Parents file not found.")
            return
        if not os.path.exists(kids):
            QMessageBox.warning(self, "File Missing", "Kids file not found.")
            return

        # Determine output path
        if self.custom_output_checkbox.isChecked():
            output = self.output_line.text()
            if not output:
                QMessageBox.warning(self, "Output Missing", "Please set a custom output file.")
                return
        else:
            # Auto-generate: same folder as kids file, add '_updated'
            kids_dir = os.path.dirname(kids)
            kids_base = os.path.basename(kids)
            name, ext = os.path.splitext(kids_base)
            output = os.path.join(kids_dir, f"{name}_updated{ext}")

        self.process_btn.setEnabled(False)
        self.log_text.clear()
        self.log_text.append(f"Output file: {output}\n")
        self.log_text.append("Starting processing...\n")

        self.worker = Worker(parents, kids, output, fee, self.month_columns)
        self.worker.log_signal.connect(self.update_log)
        self.worker.finished_signal.connect(self.processing_finished)
        self.worker.start()

    def update_log(self, msg):
        self.log_text.append(msg)

    def processing_finished(self, output_file):
        self.process_btn.setEnabled(True)
        if output_file:
            QMessageBox.information(self, "Success", f"✅ Done!\nSaved to:\n{output_file}")
        else:
            QMessageBox.critical(self, "Error", "Processing failed. Check logs.")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")  # Consistent look across OS
    window = PaymentTrackerApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()